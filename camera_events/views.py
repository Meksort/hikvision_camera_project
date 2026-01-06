"""
View для приема событий от камер Hikvision.

Этот файл теперь импортирует ViewSet'ы и функции из модульных файлов.
Основная логика вынесена в:
- utils.py - утилиты и константы
- event_processor.py - обработка событий
- viewsets/ - отдельные ViewSet'ы
"""
# Импортируем утилиты и константы
from .utils import (
    SCHEDULE_TYPE_MAP,
    WEEKDAYS_SHORT,
    EXCLUDED_DEPARTMENTS,
    get_excluded_hikvision_ids,
    ensure_aware,
    clean_id,
)

# Импортируем функции обработки событий
from .event_processor import process_single_camera_event

# Импортируем ViewSet'ы
# Пока что только DepartmentViewSet вынесен в отдельный модуль
from .viewsets.department import DepartmentViewSet

# Остальные ViewSet'ы (CameraEventViewSet, EntryExitViewSet)
# пока остаются в этом файле и будут вынесены в отдельные модули позже

# Импортируем остальные функции и классы из старого файла
# (они будут постепенно вынесены в отдельные модули)
import json
import base64
import logging
import re
from datetime import datetime, timedelta, time, date
from django.http import HttpResponse, JsonResponse, FileResponse
from django.utils import timezone
from django.conf import settings
from django.db.models import Q

# Попытка использовать zoneinfo (Python 3.9+), иначе используем настройки Django
try:
    from zoneinfo import ZoneInfo
    ALMATY_TZ = ZoneInfo('Asia/Almaty')
except ImportError:
    # Для Python < 3.9 используем настройки Django
    ALMATY_TZ = None
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import utils
from io import BytesIO

from .models import CameraEvent, EntryExit, Employee, Department
from .serializers import CameraEventSerializer, EntryExitSerializer, DepartmentSerializer
from .schedule_matcher import ScheduleMatcher
from .sql_reports import generate_round_the_clock_report_sql

logger = logging.getLogger(__name__)

# Множества для быстрой проверки колонок
RED_COLUMNS_EMPTY = {}  # Колонки, которые должны быть красными для пустых строк

# process_single_camera_event вынесена в event_processor.py
# Импортируется выше


def recalculate_entries_exits(start_date=None, end_date=None):
    """
    Пересчитывает все записи EntryExit из существующих CameraEvent.
    Использует IP адреса камер для определения входа/выхода.
    
    Args:
        start_date: Начальная дата для фильтрации (datetime). Если None, обрабатывает все события.
        end_date: Конечная дата для фильтрации (datetime). Если None, обрабатывает все события.
    """
    start_time = timezone.now()
    
    try:
        created_count = 0
        updated_count = 0
        
        # Получаем все события с hikvision_id и event_time
        try:
            events = CameraEvent.objects.filter(
                hikvision_id__isnull=False,
                event_time__isnull=False
            )
            
            # Фильтруем по датам, если указаны
            if start_date:
                # Убеждаемся, что start_date в правильном формате
                start_date = timezone.make_aware(start_date) if timezone.is_naive(start_date) else start_date
                events = events.filter(event_time__gte=start_date)
            
            if end_date:
                # Убеждаемся, что end_date в правильном формате
                end_date = timezone.make_aware(end_date) if timezone.is_naive(end_date) else end_date
                # Добавляем один день, чтобы включить весь указанный день
                end_date_with_time = end_date + timedelta(days=1)
                events = events.filter(event_time__lt=end_date_with_time)
            
            events = events.order_by('event_time')
        except Exception as e:
            logger.error(f"Ошибка при получении событий: {e}", exc_info=True)
            return {"created": 0, "updated": 0, "error": str(e)}
    
        # Группируем события по сотруднику и дате
        events_by_employee_date = {}
        events_without_type = 0
        events_with_ip = 0
        events_with_device_name = 0
        
        try:
            for event in events:
                try:
                    if not event.hikvision_id or not event.event_time:
                        continue
                    
                    # Очищаем ID от ведущих нулей для правильной группировки
                    try:
                        clean_employee_id = clean_id(event.hikvision_id)
                    except Exception as e:
                        logger.warning(f"Ошибка при очистке ID события (id={event.id if hasattr(event, 'id') else 'unknown'}): {e}")
                        continue
                    
                    # Определяем тип события (вход/выход) по IP адресу
                    camera_ip = None
                    try:
                        if event.raw_data and isinstance(event.raw_data, dict):
                            # Проверяем разные варианты структуры данных
                            outer_event = event.raw_data.get("AccessControllerEvent", {})
                            if isinstance(outer_event, dict):
                                # Проверяем, есть ли вложенный AccessControllerEvent
                                if "AccessControllerEvent" in outer_event:
                                    inner_event = outer_event["AccessControllerEvent"]
                                    if isinstance(inner_event, dict):
                                        camera_ip = (
                                            inner_event.get("ipAddress") or
                                            inner_event.get("remoteHostAddr") or
                                            inner_event.get("ip") or
                                            None
                                        )
                                
                                # Если не нашли во внутреннем, проверяем внешний
                                if not camera_ip:
                                    camera_ip = (
                                        outer_event.get("ipAddress") or
                                        outer_event.get("remoteHostAddr") or
                                        outer_event.get("ip") or
                                        None
                                    )
                            
                            # Также проверяем напрямую в raw_data
                            if not camera_ip:
                                camera_ip = (
                                    event.raw_data.get("ipAddress") or
                                    event.raw_data.get("remoteHostAddr") or
                                    event.raw_data.get("ip") or
                                    None
                                )
                    except Exception as e:
                        logger.warning(f"Ошибка при извлечении IP из события (id={event.id if hasattr(event, 'id') else 'unknown'}): {e}")
                        camera_ip = None
                    
                    # Определяем тип события
                    # ПРИОРИТЕТ: Сначала проверяем IP адрес (самый надежный способ)
                    is_entry = False
                    is_exit = False
                    
                    try:
                        if camera_ip:
                            camera_ip_str = str(camera_ip)
                            # ВЫХОД: IP содержит 143 или 192.168.1.143
                            if "192.168.1.143" in camera_ip_str or camera_ip_str.endswith(".143") or camera_ip_str == "143":
                                is_exit = True
                            # ВХОД: IP содержит 124 или 192.168.1.124
                            elif "192.168.1.124" in camera_ip_str or camera_ip_str.endswith(".124") or camera_ip_str == "124":
                                is_entry = True
                        
                        # Если IP не определен, проверяем device_name
                        if not is_entry and not is_exit:
                            device_name_lower = (event.device_name or "").lower()
                            # ВХОД: проверяем ключевые слова
                            is_entry = any(word in device_name_lower for word in ['вход', 'entry', 'входная', 'вход 1', 'вход1', '124'])
                            # ВЫХОД: проверяем ключевые слова
                            is_exit = any(word in device_name_lower for word in ['выход', 'exit', 'выходная', 'выход 1', 'выход1', '143'])
                    except Exception as e:
                        logger.warning(f"Ошибка при определении типа события (id={event.id if hasattr(event, 'id') else 'unknown'}): {e}")
                        continue
                    
                    if not (is_entry or is_exit):
                        events_without_type += 1
                        continue
                    
                    if camera_ip:
                        events_with_ip += 1
                    else:
                        events_with_device_name += 1
                    
                    try:
                        event_date = event.event_time.date()
                        # Используем очищенный ID для группировки
                        key = (clean_employee_id, event_date)
                        
                        if key not in events_by_employee_date:
                            events_by_employee_date[key] = {
                                'entry_events': [],
                                'exit_events': [],
                            }
                        
                        if is_entry:
                            events_by_employee_date[key]['entry_events'].append(event)
                        elif is_exit:
                            events_by_employee_date[key]['exit_events'].append(event)
                    except Exception as e:
                        logger.warning(f"Ошибка при группировке события (id={event.id if hasattr(event, 'id') else 'unknown'}): {e}")
                        continue
                except Exception as e:
                    logger.warning(f"Ошибка при обработке события (id={event.id if hasattr(event, 'id') else 'unknown'}): {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при группировке событий: {e}", exc_info=True)
            return {"created": 0, "updated": 0, "error": str(e)}
    
        # Обрабатываем каждую группу событий
        total_groups = len(events_by_employee_date)
        
        group_index = 0
        for (hikvision_id, event_date), events_data in events_by_employee_date.items():
            group_index += 1
            try:
                # Сортируем события по времени
                try:
                    entry_events = sorted(events_data['entry_events'], key=lambda e: e.event_time)
                    exit_events = sorted(events_data['exit_events'], key=lambda e: e.event_time)
                except Exception as e:
                    logger.warning(f"Ошибка при сортировке событий для сотрудника {hikvision_id}, дата {event_date}: {e}")
                    continue
                
                # Для ночных смен нужно также проверять выходы следующего дня
                # Получаем выходы следующего дня
                next_date = event_date + timedelta(days=1)
                next_day_key = (hikvision_id, next_date)
                next_day_exit_events = []
                if next_day_key in events_by_employee_date:
                    try:
                        next_day_exit_events = sorted(events_by_employee_date[next_day_key]['exit_events'], key=lambda e: e.event_time)
                    except Exception:
                        pass
                
                # Объединяем выходы текущего дня и следующего дня
                all_exit_events = exit_events + next_day_exit_events
                all_exit_events = sorted(all_exit_events, key=lambda e: e.event_time)
                
                # Создаем или обновляем записи EntryExit
                entry_idx = 0
                exit_idx = 0
                
                while entry_idx < len(entry_events):
                    try:
                        entry_event = entry_events[entry_idx]
                        entry_time = entry_event.event_time
                        
                        # Ищем соответствующий выход (ближайший после входа)
                        # Для ночных смен выход может быть на следующий день
                        matching_exit_event = None
                        matching_exit_idx = None
                        
                        try:
                            # УЛУЧШЕННАЯ ЛОГИКА: Если есть вход (IP 124) и выход (IP 143) за один день,
                            # обязательно находим соответствующий выход
                            
                            # Сначала ищем выходы в текущем дне (приоритет - ближайший после входа)
                            for i in range(exit_idx, len(exit_events)):
                                if exit_events[i].event_time > entry_time:
                                    # Проверяем, что это действительно выход (IP 143 или название содержит "выход")
                                    exit_event = exit_events[i]
                                    is_valid_exit = False
                                    
                                    # Проверяем IP адрес выхода
                                    exit_camera_ip = None
                                    if exit_event.raw_data and isinstance(exit_event.raw_data, dict):
                                        outer_event = exit_event.raw_data.get("AccessControllerEvent", {})
                                        if isinstance(outer_event, dict):
                                            if "AccessControllerEvent" in outer_event:
                                                inner_event = outer_event["AccessControllerEvent"]
                                                if isinstance(inner_event, dict):
                                                    exit_camera_ip = inner_event.get("ipAddress") or inner_event.get("remoteHostAddr") or inner_event.get("ip")
                                            if not exit_camera_ip:
                                                exit_camera_ip = outer_event.get("ipAddress") or outer_event.get("remoteHostAddr") or outer_event.get("ip")
                                        if not exit_camera_ip:
                                            exit_camera_ip = exit_event.raw_data.get("ipAddress") or exit_event.raw_data.get("remoteHostAddr") or exit_event.raw_data.get("ip")
                                    
                                    if exit_camera_ip:
                                        exit_ip_str = str(exit_camera_ip)
                                        if "192.168.1.143" in exit_ip_str or exit_ip_str.endswith(".143") or exit_ip_str == "143":
                                            is_valid_exit = True
                                    
                                    # Если IP не определен, проверяем device_name
                                    if not is_valid_exit:
                                        exit_device_lower = (exit_event.device_name or "").lower()
                                        is_valid_exit = any(word in exit_device_lower for word in ['выход', 'exit', 'выходная', 'выход 1', 'выход1', '143'])
                                    
                                    # Если это валидный выход, используем его
                                    if is_valid_exit:
                                        matching_exit_event = exit_event
                                        matching_exit_idx = i
                                        break
                            
                            # Если не нашли в текущем дне, ищем в следующем дне (для ночных смен)
                            if not matching_exit_event and next_day_exit_events:
                                # Для ночных смен выход должен быть не позже чем через 16 часов после входа
                                max_exit_time = entry_time + timedelta(hours=16)
                                # Минимум 30 минут после входа (чтобы исключить случайные совпадения, но не слишком строго)
                                min_exit_time = entry_time + timedelta(minutes=30)
                                for exit_event in next_day_exit_events:
                                    if exit_event.event_time > entry_time and min_exit_time <= exit_event.event_time <= max_exit_time:
                                        # Проверяем, что это действительно выход
                                        is_valid_exit = False
                                        exit_camera_ip = None
                                        if exit_event.raw_data and isinstance(exit_event.raw_data, dict):
                                            outer_event = exit_event.raw_data.get("AccessControllerEvent", {})
                                            if isinstance(outer_event, dict):
                                                if "AccessControllerEvent" in outer_event:
                                                    inner_event = outer_event["AccessControllerEvent"]
                                                    if isinstance(inner_event, dict):
                                                        exit_camera_ip = inner_event.get("ipAddress") or inner_event.get("remoteHostAddr") or inner_event.get("ip")
                                                if not exit_camera_ip:
                                                    exit_camera_ip = outer_event.get("ipAddress") or outer_event.get("remoteHostAddr") or outer_event.get("ip")
                                            if not exit_camera_ip:
                                                exit_camera_ip = exit_event.raw_data.get("ipAddress") or exit_event.raw_data.get("remoteHostAddr") or exit_event.raw_data.get("ip")
                                        
                                        if exit_camera_ip:
                                            exit_ip_str = str(exit_camera_ip)
                                            if "192.168.1.143" in exit_ip_str or exit_ip_str.endswith(".143") or exit_ip_str == "143":
                                                is_valid_exit = True
                                        
                                        if not is_valid_exit:
                                            exit_device_lower = (exit_event.device_name or "").lower()
                                            is_valid_exit = any(word in exit_device_lower for word in ['выход', 'exit', 'выходная', 'выход 1', 'выход1', '143'])
                                        
                                        if is_valid_exit:
                                            matching_exit_event = exit_event
                                            matching_exit_idx = len(exit_events) + next_day_exit_events.index(exit_event)
                                            break
                        except Exception as e:
                            logger.warning(f"Ошибка при поиске соответствующего выхода для сотрудника {hikvision_id}: {e}")
                        
                        # Проверяем, существует ли уже запись (используем очищенный ID)
                        try:
                            clean_hikvision_id = clean_id(hikvision_id)
                            existing = EntryExit.objects.filter(
                                hikvision_id=clean_hikvision_id,
                                entry_time__date=event_date,
                                entry_time=entry_time
                            ).first()
                            
                            # Если не нашли по очищенному ID, пробуем найти по исходному
                            if not existing:
                                existing = EntryExit.objects.filter(
                                    hikvision_id=hikvision_id,
                                    entry_time__date=event_date,
                                    entry_time=entry_time
                                ).first()
                        except Exception as e:
                            logger.warning(f"Ошибка при поиске существующей записи для сотрудника {hikvision_id}: {e}")
                            existing = None
                        
                        if existing:
                            # Обновляем существующую запись, если нашелся выход
                            try:
                                if matching_exit_event and not existing.exit_time:
                                    duration = matching_exit_event.event_time - existing.entry_time
                                    existing.exit_time = matching_exit_event.event_time
                                    existing.device_name_exit = matching_exit_event.device_name
                                    existing.work_duration_seconds = int(duration.total_seconds())
                                    # Обновляем ID на очищенный, если нужно
                                    if existing.hikvision_id != clean_hikvision_id:
                                        existing.hikvision_id = clean_hikvision_id
                                    existing.save()
                                    updated_count += 1
                            except Exception as e:
                                logger.warning(f"Ошибка при обновлении записи EntryExit (id={existing.id if hasattr(existing, 'id') else 'unknown'}): {e}")
                        else:
                            # Создаем новую запись
                            try:
                                work_duration_seconds = None
                                exit_time = None
                                device_name_exit = None
                                
                                if matching_exit_event:
                                    exit_time = matching_exit_event.event_time
                                    device_name_exit = matching_exit_event.device_name
                                    duration = exit_time - entry_time
                                    work_duration_seconds = int(duration.total_seconds())
                                
                                EntryExit.objects.create(
                                    hikvision_id=clean_hikvision_id,  # Сохраняем очищенный ID
                                    entry_time=entry_time,
                                    exit_time=exit_time,
                                    device_name_entry=entry_event.device_name,
                                    device_name_exit=device_name_exit,
                                    work_duration_seconds=work_duration_seconds,
                                )
                                created_count += 1
                            except Exception as e:
                                logger.warning(f"Ошибка при создании записи EntryExit для сотрудника {hikvision_id}: {e}")
                        
                        entry_idx += 1
                        if matching_exit_idx is not None and matching_exit_idx < len(exit_events):
                            # Обновляем exit_idx только если выход был из текущего дня
                            exit_idx = matching_exit_idx + 1
                    except Exception as e:
                        logger.warning(f"Ошибка при обработке записи входа для сотрудника {hikvision_id}: {e}")
                        entry_idx += 1
                        continue
            except Exception as e:
                logger.warning(f"Ошибка при обработке группы событий для сотрудника {hikvision_id}, дата {event_date}: {e}")
                continue
            
            # Выводим прогресс каждые 500 групп или для последней
            if group_index % 500 == 0 or group_index == total_groups:
                logger.info(f"Прогресс: {group_index}/{total_groups} ({group_index*100//total_groups if total_groups > 0 else 0}%)")
    
        # Выводим сообщение об успешном обновлении
        end_time = timezone.now()
        duration = (end_time - start_time).total_seconds()
        logger.info(f"[{end_time.strftime('%H:%M:%S')}] Данные: {duration:.1f}с, создано={created_count}, обновлено={updated_count}")
        
        return {"created": created_count, "updated": updated_count}
    
    except Exception as e:
        end_time = timezone.now()
        duration = (end_time - start_time).total_seconds()
        logger.error(f"[{end_time.strftime('%H:%M:%S')}] Ошибка при обновлении данных (время до ошибки: {duration:.1f}с): {e}", exc_info=True)
        return {"created": 0, "updated": 0, "error": str(e)}


class CameraEventViewSet(viewsets.ModelViewSet):
    """
    ViewSet для приема событий от камер Hikvision.
    
    Endpoint: POST /api/v1/camera-events/
    """
    queryset = CameraEvent.objects.all()
    permission_classes = [AllowAny]  # Камеры не используют аутентификацию
    serializer_class = CameraEventSerializer
    
    def create(self, request, *args, **kwargs):
        """
        Прием события от камеры Hikvision.
        
        Поддерживает форматы:
        - multipart/form-data (с event_log и Picture)
        - application/json
        """
        # Логирование входящего запроса от камеры
        client_ip = request.META.get('REMOTE_ADDR', 'unknown')
        print(f"\n{'='*60}")
        print(f"📹 ПОЛУЧЕНО СОБЫТИЕ ОТ КАМЕРЫ")
        print(f"{'='*60}")
        print(f"IP адрес камеры: {client_ip}")
        print(f"Время получения: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Content-Type: {request.content_type}")
        print(f"Method: {request.method}")
        logger.info(f"📹 Получено событие от камеры IP: {client_ip}, Content-Type: {request.content_type}")
        
        try:
            # Определяем тип контента
            content_type = request.content_type or ""
            content_type_lower = content_type.lower()
            
            event_data = None
            picture_data = None
            
            # Обработка multipart/form-data
            if "multipart/form-data" in content_type_lower:
                # Извлекаем event_log или AccessControllerEvent
                event_log_raw = request.POST.get("event_log")
                access_event_raw = request.POST.get("AccessControllerEvent")
                
                # Пробуем сначала event_log
                if event_log_raw:
                    try:
                        event_data = json.loads(event_log_raw)
                    except json.JSONDecodeError as e:
                        logger.error(f"Failed to parse event_log JSON: {e}")
                        return HttpResponse("OK", status=200)
                
                # Если event_log нет, пробуем AccessControllerEvent
                elif access_event_raw:
                    try:
                        # Пробуем распарсить как JSON
                        if isinstance(access_event_raw, str):
                            parsed_event = json.loads(access_event_raw)
                            event_data = {"AccessControllerEvent": parsed_event}
                        else:
                            # Если уже словарь или другой тип
                            event_data = {"AccessControllerEvent": access_event_raw}
                    except json.JSONDecodeError as e:
                        # Если не JSON, может быть XML или другой формат
                        logger.warning(f"AccessControllerEvent is not JSON, treating as raw data: {e}")
                        # Создаем структуру с сырыми данными
                        event_data = {"AccessControllerEvent": {"raw": str(access_event_raw)}}
                
                # Извлекаем изображение
                picture_file = request.FILES.get("Picture")
                if picture_file:
                    try:
                        picture_bytes = picture_file.read()
                        picture_data = base64.b64encode(picture_bytes).decode('utf-8')
                    except Exception as e:
                        logger.error(f"Failed to process picture: {e}")
            
            # Обработка JSON
            elif "json" in content_type_lower:
                try:
                    if hasattr(request, 'data'):
                        data = request.data
                    else:
                        data = json.loads(request.body)
                    
                    event_data = data.get("event_log") or data
                    picture_data = data.get("picData")
                except json.JSONDecodeError as e:
                    logger.error(f"Failed to parse JSON: {e}")
                    return HttpResponse("OK", status=200)
            
            # Если данных нет, пробуем извлечь из AccessControllerEvent напрямую
            if not event_data:
                logger.warning("No event data found in request")
                logger.debug(f"POST keys: {list(request.POST.keys())}")
                logger.debug(f"FILES keys: {list(request.FILES.keys())}")
                
                # Пробуем извлечь AccessControllerEvent напрямую из POST
                access_event_raw = request.POST.get("AccessControllerEvent")
                if access_event_raw:
                    try:
                        # Пробуем распарсить как JSON
                        if isinstance(access_event_raw, str):
                            try:
                                parsed = json.loads(access_event_raw)
                                event_data = {"AccessControllerEvent": parsed}
                            except json.JSONDecodeError:
                                # Может быть XML или другой формат - создаем структуру
                                event_data = {"AccessControllerEvent": {"raw": access_event_raw}}
                        else:
                            # Уже словарь или другой объект
                            event_data = {"AccessControllerEvent": access_event_raw}
                    except Exception as e:
                        logger.error(f"Error processing AccessControllerEvent: {e}")
                        import traceback
                        logger.error(traceback.format_exc())
                
                # Если все еще нет данных, возвращаем OK
                if not event_data:
                    logger.warning("⚠️  Still no event data found, returning OK to camera")
                return HttpResponse("OK", status=200)
            
            # Извлекаем данные из AccessControllerEvent если есть
            if isinstance(event_data, dict) and "AccessControllerEvent" in event_data:
                # Получаем внешний объект AccessControllerEvent
                outer_event = event_data["AccessControllerEvent"]
                
                # Получаем вложенный объект AccessControllerEvent
                if isinstance(outer_event, dict):
                    # Если AccessControllerEvent - это словарь, проверяем, есть ли внутри еще один AccessControllerEvent
                    if "AccessControllerEvent" in outer_event:
                        access_event = outer_event["AccessControllerEvent"]
                    else:
                        access_event = outer_event
                else:
                    access_event = outer_event
                
                # Извлекаем дату события (но не фильтруем по дате)
                # Все события будут сохранены с их реальной датой, события до 1 ноября будут первыми в базе
                event_date = None
                if isinstance(outer_event, dict):
                    date_time_str = outer_event.get("dateTime")
                    if date_time_str:
                        try:
                            # Парсим дату в формате ISO (например, "2025-11-24T16:54:25+05:00")
                            if 'T' in date_time_str:
                                date_part = date_time_str.split('T')[0]  # Берем только дату (YYYY-MM-DD)
                            else:
                                date_part = date_time_str.split()[0]  # Если формат другой
                            
                            event_date = datetime.strptime(date_part, "%Y-%m-%d").date()
                        except Exception as e:
                            logger.warning(f"⚠️  Could not parse date '{date_time_str}': {e}")
                
                # Проверяем тип события внутри AccessControllerEvent
                # Сохраняем только события с данными о сотрудниках, пропускаем служебные (heartBeat и т.д.)
                is_valid_event = False
                if isinstance(access_event, dict):
                    sub_event_type = access_event.get("subEventType")
                    major_event_type = access_event.get("majorEventType")
                    event_type = access_event.get("eventType")
                    event_description = access_event.get("eventDescription")
                    
                    # Пропускаем служебные события типа heartBeat
                    if event_type and isinstance(event_type, str):
                        event_type_lower = event_type.lower()
                        if "heartbeat" in event_type_lower or "heart" in event_type_lower:
                            return HttpResponse("OK", status=200)
                    
                    if event_description and isinstance(event_description, str):
                        event_desc_lower = event_description.lower()
                        if "heartbeat" in event_desc_lower or "heart" in event_desc_lower:
                            return HttpResponse("OK", status=200)
                    
                    # Проверяем наличие данных о сотруднике (ID или имя)
                    has_employee_data = (
                        access_event.get("employeeId") or
                        access_event.get("employeeID") or
                        access_event.get("employeeNo") or
                        access_event.get("employeeNoString") or
                        access_event.get("name") or
                        access_event.get("employeeName") or
                        access_event.get("employeeNameString")
                    )
                    
                    # События аутентификации по лицу (subEventType = 75)
                    if sub_event_type == 75:
                        is_valid_event = True
                        # Определяем тип события (Вход/Выход) по IP адресу
                        event_type_text = ""
                        camera_ip = None
                        if isinstance(outer_event, dict):
                            camera_ip = (
                                outer_event.get("ipAddress") or
                                outer_event.get("remoteHostAddr") or
                                outer_event.get("ip") or
                                None
                            )
                        if not camera_ip and isinstance(access_event, dict):
                            camera_ip = (
                                access_event.get("ipAddress") or
                                access_event.get("remoteHostAddr") or
                                access_event.get("ip") or
                                None
                            )
                        if camera_ip:
                            if "192.168.1.143" in str(camera_ip):
                                event_type_text = "Выход"
                            elif "192.168.1.124" in str(camera_ip):
                                event_type_text = "Вход"
                        
                        # Если IP не определен, проверяем device_name из access_event
                        if not event_type_text and isinstance(access_event, dict):
                            device_name_check = access_event.get("deviceName") or access_event.get("door") or access_event.get("doorName") or ""
                            device_name_lower = str(device_name_check).lower()
                            if any(word in device_name_lower for word in ['выход', 'exit', 'выходная', 'выход 1', 'выход1', '143']):
                                event_type_text = "Выход"
                            elif any(word in device_name_lower for word in ['вход', 'entry', 'входная', 'вход 1', 'вход1', '124']):
                                event_type_text = "Вход"
                        
                        # Если тип не определен, используем значение по умолчанию
                        if not event_type_text:
                            event_type_text = "Событие"
                        
                        # Получаем время события для логирования
                        event_time_for_log = None
                        if isinstance(outer_event, dict):
                            event_time_for_log = outer_event.get("dateTime") or access_event.get("dateTime") if isinstance(access_event, dict) else None
                        event_time_display = f" [{event_time_for_log}]" if event_time_for_log else ""
                        logger.info(f"✅ {event_type_text}{event_time_display}")
                    # События контроля доступа (majorEventType = 5) с данными о сотруднике
                    elif major_event_type == 5 and has_employee_data:
                        is_valid_event = True
                    # Любое событие с данными о сотруднике
                    elif has_employee_data:
                        is_valid_event = True
                    
                    # Если событие не содержит данных о сотруднике, пропускаем
                    if not is_valid_event:
                        return HttpResponse("OK", status=200)
                else:
                    # Если access_event не словарь, пропускаем
                    return HttpResponse("OK", status=200)
                
                # Извлекаем ID от Hikvision (пробуем разные варианты названий полей)
                hikvision_id = None
                if isinstance(access_event, dict):
                    hikvision_id = (
                        access_event.get("employeeId") or
                        access_event.get("employeeID") or
                        access_event.get("employeeNo") or 
                        access_event.get("employeeNoString") or
                        access_event.get("employee_id") or
                        access_event.get("Employee ID") or
                        access_event.get("cardNo") or
                        access_event.get("cardNumber") or
                        access_event.get("cardReaderNo") or
                        access_event.get("doorNo") or
                        access_event.get("door") or
                        (str(access_event.get("id")) if access_event.get("id") is not None else None)
                    )
                
                # Извлекаем имя сотрудника
                employee_name = None
                if isinstance(access_event, dict):
                    employee_name = (
                        access_event.get("employeeName") or
                        access_event.get("name") or
                        access_event.get("employeeNameString") or
                        access_event.get("employee_name") or
                        access_event.get("Name") or
                        None
                    )
                
                # Извлекаем номер карты
                card_no = None
                if isinstance(access_event, dict):
                    card_no = (
                        access_event.get("cardNo") or
                        access_event.get("cardNumber") or
                        access_event.get("card") or
                        access_event.get("Card No.") or
                        access_event.get("card_no") or
                        None
                    )
                
                # Извлекаем дверь/устройство
                device_name = None
                if isinstance(access_event, dict):
                    device_name = (
                        access_event.get("deviceName") or
                        access_event.get("door") or
                        access_event.get("doorName") or
                        access_event.get("doorNo") or
                        access_event.get("Door") or
                        access_event.get("device_name") or
                        None
                    )
                
                # Извлекаем время события
                # Время может быть во внешнем объекте (outer_event) или во внутреннем (access_event)
                event_time_str = None
                if isinstance(outer_event, dict):
                    event_time_str = (
                        outer_event.get("dateTime") or
                        outer_event.get("time") or
                        outer_event.get("eventTime") or
                        None
                    )
                
                # Если не нашли во внешнем, проверяем внутренний
                if not event_time_str and isinstance(access_event, dict):
                    event_time_str = (
                        access_event.get("time") or
                        access_event.get("dateTime") or
                        access_event.get("eventTime") or
                        access_event.get("Time") or
                        access_event.get("event_time") or
                        None
                    )
                
                # Если все еще не нашли, проверяем event_data
                if not event_time_str:
                    event_time_str = (
                        event_data.get("dateTime") or
                        event_data.get("time") or
                        None
                    )
                
                # Проверка имени сотрудника (обязательно для сохранения)
                # Событие уже прошло проверку по subEventType, теперь проверяем наличие имени
                if not employee_name or employee_name.strip() == "":
                    return HttpResponse("OK", status=200)
                else:
                    # Используем уже извлеченное время для логирования
                    event_time_display = f" [{event_time_str}]" if event_time_str else ""
                    logger.info(f"✅ Employee name found: '{employee_name}' - will save{event_time_display}")
                
                # Парсим event_time в datetime объект перед сохранением
                event_time_parsed = None
                if event_time_str:
                    if isinstance(event_time_str, str):
                        try:
                            # Пробуем разные форматы даты
                            for fmt in ['%Y-%m-%dT%H:%M:%S%z', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f%z', '%Y-%m-%dT%H:%M:%S.%f']:
                                try:
                                    event_time_parsed = datetime.strptime(event_time_str, fmt)
                                    if timezone.is_naive(event_time_parsed):
                                        event_time_parsed = timezone.make_aware(event_time_parsed)
                                    break
                                except ValueError:
                                    continue
                            if event_time_parsed is None:
                                # Если не удалось распарсить, используем текущее время
                                logger.warning(f"Could not parse event_time '{event_time_str}', using current time")
                                event_time_parsed = timezone.now()
                        except Exception as e:
                            logger.warning(f"Error parsing event_time '{event_time_str}': {e}, using current time")
                            event_time_parsed = timezone.now()
                    elif isinstance(event_time_str, datetime):
                        event_time_parsed = event_time_str
                        if timezone.is_naive(event_time_parsed):
                            event_time_parsed = timezone.make_aware(event_time_parsed)
                    else:
                        event_time_parsed = timezone.now()
                else:
                    event_time_parsed = timezone.now()
                
                # Если ничего не нашли, логируем все ключи
                if not hikvision_id and isinstance(access_event, dict):
                    logger.warning(f"⚠️  Could not find employee ID. Available keys: {list(access_event.keys())}")
                    logger.warning(f"⚠️  Full access_event: {json.dumps(access_event, indent=2, ensure_ascii=False)}")
            else:
                # Если нет AccessControllerEvent, проверяем event_data напрямую
                # Но для событий без AccessControllerEvent тоже нужно проверить тип
                employee_name = None  # Инициализируем переменную
                
                event_type_direct = (
                    event_data.get("eventType") or
                    event_data.get("eventTypes") or
                    event_data.get("eventDescription") or
                    event_data.get("event") or
                    None
                )
                
                # Извлекаем имя сотрудника из event_data
                employee_name_direct = (
                    event_data.get("employeeName") or
                    event_data.get("name") or
                    event_data.get("employeeNameString") or
                    event_data.get("employee_name") or
                    event_data.get("Name") or
                    None
                )
                
                # Проверяем тип события для событий без AccessControllerEvent
                should_save_direct = False
                if event_type_direct:
                    event_type_lower = str(event_type_direct).lower()
                    # Пропускаем служебные события
                    if "heartbeat" in event_type_lower or "heart" in event_type_lower:
                        return HttpResponse("OK", status=200)
                    # Проверяем наличие данных о сотруднике
                    if employee_name_direct and employee_name_direct.strip():
                        should_save_direct = True
                    else:
                        return HttpResponse("OK", status=200)
                else:
                    # Если тип не определен, проверяем наличие данных о сотруднике
                    if employee_name_direct and employee_name_direct.strip():
                        should_save_direct = True
                    else:
                        # Проверяем, не является ли это служебным событием
                        if isinstance(event_data, dict):
                            for key, value in event_data.items():
                                if isinstance(value, str):
                                    value_lower = value.lower()
                                    if "heartbeat" in value_lower or "heart" in value_lower:
                                        return HttpResponse("OK", status=200)
                        return HttpResponse("OK", status=200)
                
                # Сохраняем имя сотрудника
                if should_save_direct and employee_name_direct and employee_name_direct.strip():
                    employee_name = employee_name_direct
                
                # Пробуем извлечь ID из разных возможных полей
                hikvision_id = (
                    event_data.get("cardNo") or 
                    event_data.get("employeeNo") or 
                    event_data.get("employeeNoString") or
                    event_data.get("cardReaderNo") or
                    event_data.get("doorNo") or
                    str(event_data.get("id", ""))
                )
                device_name = event_data.get("deviceName")
                event_time_str = event_data.get("dateTime")
            
            # Парсим event_time в datetime объект перед сохранением
            event_time_parsed = None
            if event_time_str:
                if isinstance(event_time_str, str):
                    try:
                        # Пробуем разные форматы даты
                        for fmt in ['%Y-%m-%dT%H:%M:%S%z', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f%z', '%Y-%m-%dT%H:%M:%S.%f']:
                            try:
                                event_time_parsed = datetime.strptime(event_time_str, fmt)
                                if timezone.is_naive(event_time_parsed):
                                    event_time_parsed = timezone.make_aware(event_time_parsed)
                                break
                            except ValueError:
                                continue
                        if event_time_parsed is None:
                            # Если не удалось распарсить, используем текущее время
                            logger.warning(f"Could not parse event_time '{event_time_str}', using current time")
                            event_time_parsed = timezone.now()
                    except Exception as e:
                        logger.warning(f"Error parsing event_time '{event_time_str}': {e}, using current time")
                        event_time_parsed = timezone.now()
                elif isinstance(event_time_str, datetime):
                    event_time_parsed = event_time_str
                    if timezone.is_naive(event_time_parsed):
                        event_time_parsed = timezone.make_aware(event_time_parsed)
                else:
                    event_time_parsed = timezone.now()
            else:
                event_time_parsed = timezone.now()
            
            # Создаем запись события
            try:
                camera_event = CameraEvent.objects.create(
                    hikvision_id=hikvision_id,
                    device_name=device_name,
                    event_time=event_time_parsed,  # Сохраняем как datetime объект
                    picture_data=picture_data,
                    raw_data=event_data,  # Все данные сохраняются здесь для доступа к employeeName, cardNo, eventType и т.д.
                )
                
                # Определяем тип события для логирования
                event_type_str = "N/A"
                employee_name_display = employee_name if 'employee_name' in locals() and employee_name else "N/A"
                try:
                    # Пробуем получить тип события из access_event, если он доступен
                    if 'access_event' in locals() and isinstance(access_event, dict):
                        sub_event_type = access_event.get("subEventType")
                        major_event_type = access_event.get("majorEventType")
                        if sub_event_type == 75:
                            event_type_str = "Authenticated via Face (subEventType=75)"
                        elif sub_event_type:
                            event_type_str = f"subEventType={sub_event_type}, majorEventType={major_event_type}"
                except:
                    pass
                
                # Логирование успешного сохранения события
                print(f"✅ СОБЫТИЕ СОХРАНЕНО:")
                print(f"   ID сотрудника: {hikvision_id or 'N/A'}")
                print(f"   Имя: {employee_name_display}")
                print(f"   Устройство: {device_name or 'N/A'}")
                print(f"   Время события: {event_time_parsed.strftime('%Y-%m-%d %H:%M:%S') if event_time_parsed else 'N/A'}")
                print(f"   Тип события: {event_type_str}")
                print(f"   CameraEvent ID: {camera_event.id}")
                print(f"{'='*60}\n")
                
                logger.info(f"✅ Событие сохранено: ID={hikvision_id}, Имя={employee_name_display}, Устройство={device_name}, Время={event_time_parsed}, CameraEvent ID={camera_event.id}")
                
            except Exception as e:
                print(f"❌ ОШИБКА при сохранении события: {e}")
                logger.error(f"Error creating CameraEvent: {e}", exc_info=True)
                logger.error(f"Event data: hikvision_id={hikvision_id}, device_name={device_name}, event_time={event_time_parsed}")
                return HttpResponse("OK", status=200)
            
            # Если данные не извлечены, логируем для анализа
            if (not hikvision_id or hikvision_id == "") and (not employee_name or employee_name == ""):
                logger.warning(f"⚠️  No data extracted from event ID={camera_event.id}!")
                logger.warning(f"⚠️  event_data type: {type(event_data)}")
                if isinstance(event_data, dict):
                    logger.warning(f"⚠️  event_data keys: {list(event_data.keys())}")
                    if "AccessControllerEvent" in event_data:
                        access_event = event_data["AccessControllerEvent"]
                        logger.warning(f"⚠️  AccessControllerEvent type: {type(access_event)}")
                        if isinstance(access_event, dict):
                            logger.warning(f"⚠️  AccessControllerEvent keys: {list(access_event.keys())}")
                            logger.warning(f"⚠️  AccessControllerEvent content: {json.dumps(access_event, indent=2, ensure_ascii=False)}")
                        elif isinstance(access_event, str):
                            logger.warning(f"⚠️  AccessControllerEvent is string: {access_event}")
                        else:
                            logger.warning(f"⚠️  AccessControllerEvent value: {access_event}")
            
            return HttpResponse("OK", status=200)
            
        except Exception as e:
            logger.exception(f"Error processing camera event: {e}")
            # Всегда возвращаем OK камере, чтобы она не повторяла запрос
            return HttpResponse("OK", status=200)
    
    def list(self, request, *args, **kwargs):
        """Список всех событий."""
        queryset = self.get_queryset()
        
        # Фильтры
        hikvision_id = request.query_params.get("hikvision_id")
        device_name = request.query_params.get("device_name")
        
        if hikvision_id:
            queryset = queryset.filter(hikvision_id__icontains=hikvision_id)
        if device_name:
            queryset = queryset.filter(device_name__icontains=device_name)
        
        # Пагинация
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Экспорт событий в Excel формат.
        Использует EntryExit как основной источник данных.
        
        Параметры фильтрации:
        - hikvision_id - фильтр по ID от Hikvision
        - device_name - фильтр по названию устройства
        - start_date - начальная дата (формат: YYYY-MM-DD)
        - end_date - конечная дата (формат: YYYY-MM-DD)
        - use_sql - использовать оптимизированные SQL запросы (true/false, по умолчанию false)
        """
        # Фильтры
        hikvision_id = request.query_params.get("hikvision_id")
        device_name = request.query_params.get("device_name")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        use_sql = request.query_params.get("use_sql", "false").lower() == "true"
        
        # ВСЕГДА используем оптимизированные SQL запросы для максимальной производительности
        return self._export_excel_sql(hikvision_id, device_name, start_date, end_date)
    
    def _export_excel_sql(self, hikvision_id, device_name, start_date, end_date):
        """
        Экспорт отчетов (один лист) с использованием
        SQL-функции generate_comprehensive_attendance_report_sql.
        """
        from .sql_reports import generate_comprehensive_attendance_report_sql
        from .utils import get_excluded_hikvision_ids
        
        # Получаем исключаемые ID
        excluded_hikvision_ids = get_excluded_hikvision_ids()
        
        # Получаем данные через SQL функцию
        results, start_date_obj, end_date_obj = generate_comprehensive_attendance_report_sql(
            hikvision_id=hikvision_id,
            start_date=start_date,
            end_date=end_date,
            device_name=device_name,
            excluded_hikvision_ids=excluded_hikvision_ids
        )
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "События камер"
        
        # Заголовки
        headers = ["Имя", "Подразделение", "Дата", "День недели", "Тип графика", 
                   "Время графика", "Время входа", "Время выхода", "Продолжительность работы", 
                   "Опоздание", "Ранний уход", "Ранний приход", "Поздний выход"]
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Группируем результаты по сотруднику и дате
        data_by_employee_date = {}
        for result in results:
            emp_id = result.get('hikvision_id')
            report_date = result.get('report_date')
            if isinstance(report_date, date):
                date_key = report_date
            elif isinstance(report_date, str):
                try:
                    date_key = datetime.strptime(report_date, "%Y-%m-%d").date()
                except:
                    continue
            else:
                continue
            
            key = (emp_id, date_key)
            if key not in data_by_employee_date:
                data_by_employee_date[key] = result
        
        # Получаем информацию о сотрудниках для пустых строк
        unique_employee_ids = set()
        for result in results:
            unique_employee_ids.add(result.get('hikvision_id'))
        
        if hikvision_id:
            clean_hikvision_id = clean_id(hikvision_id)
            unique_employee_ids.add(clean_hikvision_id)
        
        # Загружаем информацию о сотрудниках и графиках
        employee_info_cache = {}
        schedule_cache = {}
        
        if unique_employee_ids:
            employees = Employee.objects.only(
                'id', 'hikvision_id', 'name', 'department_id', 'department_old'
            ).filter(
                hikvision_id__in=unique_employee_ids
            ).select_related('department').prefetch_related('work_schedules')
            
            for employee in employees:
                clean_emp_id = clean_id(employee.hikvision_id)
                employee_info_cache[clean_emp_id] = {
                    'name': employee.name.replace('\n', ' ').replace('\r', ' ').strip() if employee.name else '',
                    'department': employee.department.get_full_path() if employee.department else (employee.department_old or '')
                }
                schedule = employee.work_schedules.first()
                if schedule:
                    schedule_cache[clean_emp_id] = schedule
        
        # Определяем основной employee_id для пустых строк
        main_employee_id = None
        if hikvision_id:
            main_employee_id = clean_id(hikvision_id)
        elif unique_employee_ids:
            main_employee_id = list(unique_employee_ids)[0]
        
        main_employee_info = employee_info_cache.get(main_employee_id, {})
        main_schedule = schedule_cache.get(main_employee_id)
        
        # Генерируем все даты в диапазоне
        current_date = start_date_obj
        row_num = 2
        total_duration_hours = 0.0
        total_scheduled_hours = 0.0
        
        while current_date <= end_date_obj:
            # Ищем данные для текущей даты
            found_data = None
            if main_employee_id:
                key = (main_employee_id, current_date)
                found_data = data_by_employee_date.get(key)
            
            if not found_data:
                # Пустая строка
                date_str = current_date.strftime("%d/%m/%Y")
                weekday_str = WEEKDAYS_SHORT[current_date.weekday()]
                
                ws.append([
                    main_employee_info.get('name', ''),
                    main_employee_info.get('department', ''),
                    date_str,
                    weekday_str,
                    "",
                    "Выходной",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ])
                
                # Применяем границы и стили ко всем ячейкам строки
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Красные ячейки для пустых значений
                for col_idx in [7, 8, 9]:  # G, H, I - время входа, выхода, продолжительность
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.fill = red_fill
                
                row_num += 1
                current_date += timedelta(days=1)
                continue
            
            # Обрабатываем данные
            employee_name = found_data.get('employee_name', '')
            department_name = found_data.get('department_name', '')
            report_date = found_data.get('report_date')
            # Преобразуем день недели из PostgreSQL DOW (0=воскресенье) в индекс для WEEKDAYS_SHORT (0=понедельник)
            day_of_week_sql = int(found_data.get('day_of_week', 0))
            day_of_week = (day_of_week_sql + 6) % 7  # Преобразование: 0(ВС)->6, 1(ПН)->0, 2(ВТ)->1, и т.д.
            schedule_type = found_data.get('schedule_type')
            schedule_start_time = found_data.get('schedule_start_time')
            schedule_end_time = found_data.get('schedule_end_time')
            first_entry_raw = found_data.get('first_entry')
            last_exit_raw = found_data.get('last_exit')
            total_duration_seconds = found_data.get('total_duration_seconds', 0) or 0
            late_minutes = found_data.get('late_minutes', 0) or 0
            early_leave_minutes = found_data.get('early_leave_minutes', 0) or 0
            early_arrival_minutes = found_data.get('early_arrival_minutes', 0) or 0
            late_departure_minutes = found_data.get('late_departure_minutes', 0) or 0
            
            # КОРРЕКТИРУЕМ вход и выход согласно графику работы
            # Получаем график для сотрудника
            emp_schedule = schedule_cache.get(main_employee_id) if main_employee_id else None
            
            entry_time = None
            exit_time = None
            corrected_duration = 0
            
            # ВАЖНО: для круглосуточных графиков ИСПОЛЬЗУЕМ значения,
            # уже рассчитанные в SQL (first_entry_raw / last_exit_raw),
            # чтобы сохранить логику окна 07:00–10:00 и правильного выхода.
            if schedule_type == 'round_the_clock':
                entry_time = first_entry_raw
                exit_time = last_exit_raw
            elif emp_schedule and current_date:
                # Получаем запланированное время работы по графику
                from .schedule_matcher import ScheduleMatcher
                scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(emp_schedule, current_date)
                
                if scheduled_times:
                    scheduled_start, scheduled_end = scheduled_times
                    
                    # Получаем все записи EntryExit для этого сотрудника за период
                    clean_emp_id = main_employee_id
                    if clean_emp_id:
                        # Расширяем период поиска: начинаем с 6 часов до начала смены и заканчиваем через 6 часов после окончания
                        search_start = scheduled_start - timedelta(hours=6)
                        search_end = scheduled_end + timedelta(hours=6)
                        
                        entry_exits_for_day = EntryExit.objects.filter(
                            hikvision_id=clean_emp_id,
                            entry_time__isnull=False,
                            exit_time__isnull=False
                        ).filter(
                            entry_time__gte=search_start,
                            entry_time__lte=search_end
                        ).order_by('entry_time')
                        
                        # Находим записи, которые попадают в период графика
                        valid_entries = []
                        valid_exits = []
                        
                        for ee in entry_exits_for_day:
                            if not ee.entry_time or not ee.exit_time:
                                continue
                            
                            entry_aware = ensure_aware(ee.entry_time)
                            exit_aware = ensure_aware(ee.exit_time)
                            
                            # Проверяем, попадает ли запись в период графика
                            # Вход должен быть близок к началу смены (не раньше чем за 2 часа до начала, не позже чем через 2 часа после начала)
                            # Выход должен быть близок к окончанию смены (не раньше чем за 2 часа до окончания, не позже чем через 2 часа после окончания)
                            
                            entry_to_start_diff = abs((entry_aware - scheduled_start).total_seconds())
                            exit_to_end_diff = abs((exit_aware - scheduled_end).total_seconds())
                            
                            # Для ночных смен (когда scheduled_end > scheduled_start + 12 часов)
                            if scheduled_end > scheduled_start + timedelta(hours=12):
                                # Выход может быть на следующий день
                                if exit_aware < scheduled_start:
                                    # Это выход предыдущей смены, пропускаем
                                    continue
                            
                            # Проверяем, что вход и выход логичны (вход до выхода)
                            if entry_aware >= exit_aware:
                                continue
                            
                            # Проверяем, что запись попадает в разумные границы относительно графика
                            # Вход должен быть в пределах от (начало смены - 2 часа) до (начало смены + 4 часа)
                            # Выход должен быть в пределах от (конец смены - 4 часа) до (конец смены + 2 часа)
                            entry_min = scheduled_start - timedelta(hours=2)
                            entry_max = scheduled_start + timedelta(hours=4)
                            exit_min = scheduled_end - timedelta(hours=4)
                            exit_max = scheduled_end + timedelta(hours=2)
                            
                            if entry_min <= entry_aware <= entry_max:
                                valid_entries.append((entry_aware, exit_aware))
                            
                            if exit_min <= exit_aware <= exit_max:
                                valid_exits.append((entry_aware, exit_aware))
                        
                        # Выбираем лучшую запись - ту, где вход ближе всего к началу смены
                        if valid_entries:
                            # Сортируем по близости входа к началу смены
                            valid_entries.sort(key=lambda x: abs((x[0] - scheduled_start).total_seconds()))
                            entry_time, exit_time = valid_entries[0]
                        elif valid_exits:
                            # Если нет подходящих входов, используем выходы
                            valid_exits.sort(key=lambda x: abs((x[1] - scheduled_end).total_seconds()))
                            entry_time, exit_time = valid_exits[0]
                        else:
                            # Если ничего не найдено, используем исходные значения из SQL
                            entry_time = first_entry_raw
                            exit_time = last_exit_raw
                        
                        # Рассчитываем продолжительность:
                        # - для круглосуточных графиков полагаемся на значения, рассчитанные SQL/Python-отчетами
                        #   (MIN(entry_local) и MAX(exit_local) для одной непрерывной смены, без ограничений)
                        # - для обычных графиков пересчитываем продолжительность по entry/exit для строки
                        if entry_time and exit_time and schedule_type != 'round_the_clock':
                            corrected_duration = int((exit_time - entry_time).total_seconds())
                            # Без искусственного ограничения 16 часами — оставляем фактическую длительность.
                            total_duration_seconds = corrected_duration
                else:
                    # График не определен для этого дня, используем исходные значения
                    entry_time = first_entry_raw
                    exit_time = last_exit_raw
            else:
                # Нет графика, используем исходные значения
                entry_time = first_entry_raw
                exit_time = last_exit_raw
            
            first_entry = entry_time
            last_exit = exit_time
            
            # Форматируем дату
            if isinstance(report_date, date):
                date_str = report_date.strftime("%d/%m/%Y")
            else:
                date_str = str(report_date) if report_date else current_date.strftime("%d/%m/%Y")
            
            weekday_str = WEEKDAYS_SHORT[day_of_week] if day_of_week < len(WEEKDAYS_SHORT) else ""
            
            # Тип графика
            schedule_type_str = SCHEDULE_TYPE_MAP.get(schedule_type, '')
            
            # Время графика
            schedule_time_str = ""
            if schedule_type == 'round_the_clock':
                if schedule_start_time:
                    if isinstance(schedule_start_time, time):
                        schedule_time_str = f"{schedule_start_time.strftime('%H:%M')}-{schedule_start_time.strftime('%H:%M')}"
                    else:
                        schedule_time_str = "Круглосуточно"
                else:
                    schedule_time_str = "09:00-09:00"
            elif schedule_start_time and schedule_end_time:
                if isinstance(schedule_start_time, time) and isinstance(schedule_end_time, time):
                    schedule_time_str = f"{schedule_start_time.strftime('%H:%M')}-{schedule_end_time.strftime('%H:%M')}"
            
            # Время входа и выхода
            entry_time_str = ""
            exit_time_str = ""
            
            if first_entry:
                if isinstance(first_entry, datetime):
                    entry_time_aware = ensure_aware(first_entry)
                    entry_time_local = timezone.localtime(entry_time_aware)
                    # Для круглосуточных графиков: вход всегда только время
                    if schedule_type == 'round_the_clock':
                        entry_time_str = entry_time_local.strftime("%H:%M:%S")
                    else:
                        entry_time_str = entry_time_local.strftime("%H:%M:%S")
                else:
                    entry_time_str = str(first_entry)
            
            if last_exit:
                if isinstance(last_exit, datetime):
                    exit_time_aware = ensure_aware(last_exit)
                    exit_time_local = timezone.localtime(exit_time_aware)
                    # Для круглосуточных графиков: если выход на следующий день, показываем с датой
                    if schedule_type == 'round_the_clock':
                        exit_date = exit_time_local.date()
                        # Определяем дату отчета (дату строки)
                        if isinstance(report_date, date):
                            report_date_obj = report_date
                        elif isinstance(report_date, str):
                            try:
                                report_date_obj = datetime.strptime(report_date, "%d/%m/%Y").date()
                            except:
                                report_date_obj = exit_date
                        else:
                            report_date_obj = exit_date
                        if exit_date > report_date_obj:
                            # Выход на следующий день - показываем с полной датой
                            exit_time_str = exit_time_local.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            # Выход в тот же день - только время
                            exit_time_str = exit_time_local.strftime("%H:%M:%S")
                    else:
                        exit_time_str = exit_time_local.strftime("%H:%M:%S")
                else:
                    exit_time_str = str(last_exit)
            
            # Продолжительность
            duration_str = ""
            if total_duration_seconds > 0:
                hours = int(total_duration_seconds) // 3600
                minutes = (int(total_duration_seconds) % 3600) // 60
                duration_str = f"{hours}ч {minutes}м"
                total_duration_hours += hours + (minutes / 60.0)
            
            # Форматируем опоздание и ранний уход из SQL-результатов
            late_str = ""
            if late_minutes and late_minutes > 0:
                late_hours = int(late_minutes) // 60
                late_mins = int(late_minutes) % 60
                if late_hours > 0:
                    late_str = f"{late_hours}ч {late_mins}м"
                else:
                    late_str = f"{late_mins}м"
            
            early_leave_str = ""
            if early_leave_minutes and early_leave_minutes > 0:
                early_hours = int(early_leave_minutes) // 60
                early_mins = int(early_leave_minutes) % 60
                if early_hours > 0:
                    early_leave_str = f"{early_hours}ч {early_mins}м"
                else:
                    early_leave_str = f"{early_mins}м"
            
            # Форматируем ранний приход и поздний выход
            early_arrival_str = ""
            if early_arrival_minutes and early_arrival_minutes > 0:
                early_arr_hours = int(early_arrival_minutes) // 60
                early_arr_mins = int(early_arrival_minutes) % 60
                if early_arr_hours > 0:
                    early_arrival_str = f"{early_arr_hours}ч {early_arr_mins}м"
                else:
                    early_arrival_str = f"{early_arr_mins}м"
            
            late_departure_str = ""
            if late_departure_minutes and late_departure_minutes > 0:
                late_dep_hours = int(late_departure_minutes) // 60
                late_dep_mins = int(late_departure_minutes) % 60
                if late_dep_hours > 0:
                    late_departure_str = f"{late_dep_hours}ч {late_dep_mins}м"
                else:
                    late_departure_str = f"{late_dep_mins}м"
            
            ws.append([
                employee_name,
                department_name,
                date_str,
                weekday_str,
                schedule_type_str,
                schedule_time_str,
                entry_time_str,
                exit_time_str,
                duration_str,
                late_str,
                early_leave_str,
                early_arrival_str,
                late_departure_str,
            ])
            
            # Применяем стили
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Красный цвет для опозданий, ранних уходов, ранних приходов и поздних выходов
                if col_idx == 10:  # Колонка K (Опоздание)
                    if late_str:
                        cell.fill = red_fill
                elif col_idx == 11:  # Колонка L (Ранний уход)
                    if early_leave_str:
                        cell.fill = red_fill
                elif col_idx == 12:  # Колонка M (Ранний приход)
                    if early_arrival_str:
                        cell.fill = green_fill
                elif col_idx == 13:  # Колонка N (Поздний выход)
                    if late_departure_str:
                        cell.fill = green_fill
            
            row_num += 1
            current_date += timedelta(days=1)
        
        # Автоподбор ширины колонок
        column_widths = {
            "A": 25, "B": 30, "C": 12, "D": 12, "E": 18, "F": 20,
            "G": 15, "H": 15, "I": 20, "J": 15, "K": 15, "L": 15, "M": 15, "N": 15
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        ws.row_dimensions[1].height = 30
        
        # Применяем границы ко всем ячейкам листа (на случай, если какие-то ячейки пропущены)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.border is None or not hasattr(cell.border, 'left'):
                    cell.border = border
        
        # Применяем границы ко всем ячейкам листа (на случай, если какие-то ячейки пропущены)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.border is None or not hasattr(cell.border, 'left'):
                    cell.border = border
        
        # Итоговая строка
        total_row = ws.max_row + 1
        total_hours = int(total_duration_hours)
        total_minutes = int((total_duration_hours - total_hours) * 60)
        total_duration_str = f"{total_hours}ч {total_minutes}м"
        
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_font = Font(bold=True, size=12)
        
        ws.cell(row=total_row, column=1).value = "ИТОГО:"
        ws.cell(row=total_row, column=9).value = total_duration_str
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
        
        # Сохраняем файл
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Имя файла
        unique_employee_names = set()
        if main_employee_info.get('name'):
            unique_employee_names.add(main_employee_info['name'])
        
        if len(unique_employee_names) == 1:
            employee_name_for_file = list(unique_employee_names)[0].strip()
            employee_name_for_file = re.sub(r'[<>:"/\\|?*]', '_', employee_name_for_file)
            employee_name_for_file = employee_name_for_file.replace(' ', '_')
            employee_name_for_file = re.sub(r'_+', '_', employee_name_for_file)
            
            if start_date_obj and end_date_obj:
                start_date_str = start_date_obj.strftime('%d-%m-%Y')
                end_date_str = end_date_obj.strftime('%d-%m-%Y')
                filename = f"{employee_name_for_file}_с_{start_date_str}_по_{end_date_str}.xlsx"
            else:
                filename = f"{employee_name_for_file}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"camera_events_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = FileResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        from urllib.parse import quote
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(filename)}'
        return response
    
    @action(detail=False, methods=["get"], url_path="health")
    def health_check(self, request):
        """Проверка работоспособности endpoint."""
        return JsonResponse({
            "status": "ok",
            "message": "Camera events endpoint is working",
            "endpoint": "/api/v1/camera-events/"
        })
    
    @action(detail=False, methods=["get"], url_path="latest-update")
    def latest_update(self, request):
        """
        Возвращает время последнего обновления данных с камер.
        Используется для проверки наличия новых данных.
        """
        # Получаем последнее событие
        latest_event = CameraEvent.objects.order_by('-event_time', '-created_at').first()
        
        # Получаем последнее обновление EntryExit
        latest_entry_exit = EntryExit.objects.order_by('-updated_at', '-created_at').first()
        
        # Определяем самое последнее время обновления
        latest_time = None
        if latest_event and latest_event.event_time:
            latest_time = latest_event.event_time
        if latest_entry_exit and latest_entry_exit.updated_at:
            if not latest_time or latest_entry_exit.updated_at > latest_time:
                latest_time = latest_entry_exit.updated_at
        
        # Подсчитываем количество событий за последний час
        one_hour_ago = timezone.now() - timedelta(hours=1)
        recent_events_count = CameraEvent.objects.filter(
            created_at__gte=one_hour_ago
        ).count()
        
        return JsonResponse({
            "latest_update": latest_time.isoformat() if latest_time else None,
            "recent_events_count": recent_events_count,
            "total_events": CameraEvent.objects.count(),
            "total_entry_exits": EntryExit.objects.count(),
        })
    
    @action(detail=False, methods=["post"], url_path="recalculate")
    def recalculate_entries_exits(self, request):
        """
        Пересчитывает все записи EntryExit из существующих CameraEvent.
        Полезно для обновления данных после изменений в логике или восстановления данных.
        
        Параметры (в теле запроса или query params):
        - start_date: Начальная дата в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS
        - end_date: Конечная дата в формате YYYY-MM-DD или YYYY-MM-DD HH:MM:SS
        """
        try:
            # Получаем параметры из тела запроса или query params
            if hasattr(request, 'data') and request.data:
                start_date_str = request.data.get('start_date')
                end_date_str = request.data.get('end_date')
            else:
                start_date_str = request.query_params.get('start_date')
                end_date_str = request.query_params.get('end_date')
            
            start_date = None
            end_date = None
            
            # Парсим даты
            if start_date_str:
                try:
                    if ' ' in start_date_str or 'T' in start_date_str:
                        start_date = datetime.strptime(start_date_str.replace('T', ' '), "%Y-%m-%d %H:%M:%S")
                    else:
                        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                        # Начинаем с начала дня
                        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                except ValueError as e:
                    logger.error(f"Invalid start_date format: {start_date_str}")
                    return JsonResponse({
                        "status": "error",
                        "message": f"Invalid start_date format: {e}"
                    }, status=400)
            
            if end_date_str:
                try:
                    if ' ' in end_date_str or 'T' in end_date_str:
                        end_date = datetime.strptime(end_date_str.replace('T', ' '), "%Y-%m-%d %H:%M:%S")
                    else:
                        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                        # Заканчиваем в конце дня
                        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                except ValueError as e:
                    logger.error(f"Invalid end_date format: {end_date_str}")
                    return JsonResponse({
                        "status": "error",
                        "message": f"Invalid end_date format: {e}"
                    }, status=400)
            
            # Если даты не указаны, используем значения по умолчанию: с 1 декабря по сегодня
            if not start_date:
                today = timezone.now().date()
                start_date = datetime.combine(datetime(today.year, 12, 1).date(), datetime.min.time())
                start_date = timezone.make_aware(start_date)
                logger.info(f"Using default start_date: {start_date}")
            
            if not end_date:
                end_date = timezone.now()
                logger.info(f"Using default end_date: {end_date}")
            
            # Делаем end_date aware, если нужно
            if end_date and timezone.is_naive(end_date):
                end_date = timezone.make_aware(end_date)
            
            # Делаем start_date aware, если нужно
            if start_date and timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date)
            
            result = recalculate_entries_exits(start_date=start_date, end_date=end_date)
            return JsonResponse({
                "status": "success",
                "message": "EntryExit records recalculated successfully",
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
                "created": result.get("created", 0),
                "updated": result.get("updated", 0),
            })
        except Exception as e:
            logger.error(f"Error recalculating entries_exits: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return JsonResponse({
                "status": "error",
                "message": str(e)
            }, status=500)


class EntryExitViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet для просмотра записей входов и выходов.
    Поддерживает фильтрацию по датам и экспорт в Excel.
    """
    queryset = EntryExit.objects.all()
    permission_classes = [AllowAny]
    serializer_class = EntryExitSerializer
    
    def get_queryset(self):
        """Фильтрация по параметрам запроса."""
        queryset = EntryExit.objects.all()
        
        # Исключаем сотрудников из указанных подразделений
        # Используем кэшированный список для оптимизации
        excluded_hikvision_ids = get_excluded_hikvision_ids()
        
        # Исключаем EntryExit записи с этими hikvision_id
        if excluded_hikvision_ids:
            queryset = queryset.exclude(hikvision_id__in=excluded_hikvision_ids)
        
        hikvision_id = self.request.query_params.get("hikvision_id")
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        
        if hikvision_id:
            queryset = queryset.filter(hikvision_id__icontains=hikvision_id)
        if start_date:
            # Начальная дата: с начала дня (00:00:00)
            try:
                # Пробуем распарсить как дату с временем или только дату
                if ' ' in start_date or 'T' in start_date:
                    # Уже есть время
                    start_date_clean = start_date.replace('T', ' ')
                    start_datetime = datetime.strptime(start_date_clean, "%Y-%m-%d %H:%M:%S")
                else:
                    # Только дата
                    start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
                
                # Делаем datetime aware, если он naive
                if timezone.is_naive(start_datetime):
                    start_datetime = timezone.make_aware(start_datetime)
                
                queryset = queryset.filter(entry_time__gte=start_datetime)
            except ValueError:
                # Если не удалось распарсить, используем как есть
                queryset = queryset.filter(entry_time__gte=start_date)
        if end_date:
            # Конечная дата: до конца дня (23:59:59)
            try:
                # Пробуем распарсить как дату с временем или только дату
                if ' ' in end_date or 'T' in end_date:
                    # Уже есть время
                    end_date_clean = end_date.replace('T', ' ')
                    end_datetime = datetime.strptime(end_date_clean, "%Y-%m-%d %H:%M:%S")
                else:
                    # Только дата - добавляем время до конца дня
                    end_datetime = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
                
                # Делаем datetime aware, если он naive
                if timezone.is_naive(end_datetime):
                    end_datetime = timezone.make_aware(end_datetime)
                
                queryset = queryset.filter(entry_time__lte=end_datetime)
            except ValueError:
                # Если не удалось распарсить, используем как есть
                queryset = queryset.filter(entry_time__lte=end_date)
        
        return queryset.order_by('-entry_time')
    
    @action(detail=False, methods=["get"], url_path="employees-list")
    def employees_list(self, request):
        """
        Возвращает список всех сотрудников для выпадающего списка.
        """
        employees = Employee.objects.filter(
            hikvision_id__isnull=False
        ).exclude(
            hikvision_id__in=get_excluded_hikvision_ids()
        ).select_related('department').order_by('name')
        
        employees_data = []
        for emp in employees:
            department_name = ""
            if emp.department:
                full_path = emp.department.get_full_path()
                # Убираем "АУП" или "АУП > " из начала пути
                if full_path.startswith("АУП > "):
                    department_name = full_path[6:]
                elif full_path.startswith("АУП"):
                    department_name = full_path[3:].lstrip(" > ")
                else:
                    department_name = full_path
                department_name = department_name.lstrip("/ > ")
            elif emp.department_old:
                dept_old = emp.department_old
                if dept_old.startswith("АУП/"):
                    department_name = dept_old[4:]
                elif dept_old.startswith("АУП"):
                    department_name = dept_old[3:].lstrip("/")
                else:
                    department_name = dept_old
                department_name = department_name.replace("/", " > ")
                department_name = department_name.lstrip("/ > ")
            
            employees_data.append({
                'id': emp.hikvision_id,
                'name': emp.name or '',
                'department': department_name,
                'position': emp.position or ''
            })
        
        return Response(employees_data)
    
    @action(detail=False, methods=["get"], url_path="check-date")
    def check_date(self, request):
        """
        Проверяет все данные за указанную дату.
        Показывает CameraEvent и EntryExit записи, помогает найти проблемы.
        
        Параметры:
        - date - дата в формате YYYY-MM-DD (по умолчанию: 2025-12-22)
        """
        from collections import defaultdict
        
        date_str = request.query_params.get("date", "2025-12-22")
        
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({
                "error": "Неверный формат даты. Используйте YYYY-MM-DD"
            }, status=400)
        
        start_datetime = timezone.make_aware(datetime.combine(target_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(target_date + timedelta(days=1), datetime.min.time()))
        
        result = {
            "date": date_str,
            "camera_events": {},
            "entry_exits": {},
            "problems": []
        }
        
        # 1. Получаем все CameraEvent
        camera_events = CameraEvent.objects.filter(
            event_time__gte=start_datetime,
            event_time__lt=end_datetime,
            hikvision_id__isnull=False
        ).order_by('hikvision_id', 'event_time')
        
        events_by_employee = defaultdict(list)
        for event in camera_events:
            events_by_employee[event.hikvision_id].append({
                "id": event.id,
                "time": timezone.localtime(event.event_time).strftime("%H:%M:%S"),
                "device_name": event.device_name or "",
                "raw_data": event.raw_data
            })
        
        for hikvision_id, events in events_by_employee.items():
            employee = Employee.objects.filter(hikvision_id=hikvision_id).first()
            employee_name = employee.name if employee else f"ID_{hikvision_id}"
            result["camera_events"][hikvision_id] = {
                "employee_name": employee_name,
                "events_count": len(events),
                "events": events
            }
        
        # 2. Получаем все EntryExit
        entry_exits = EntryExit.objects.filter(
            Q(entry_time__gte=start_datetime, entry_time__lt=end_datetime) |
            Q(exit_time__gte=start_datetime, exit_time__lt=end_datetime)
        ).order_by('hikvision_id', 'entry_time')
        
        entry_exits_by_employee = defaultdict(list)
        for entry_exit in entry_exits:
            entry_exits_by_employee[entry_exit.hikvision_id].append({
                "id": entry_exit.id,
                "entry_time": timezone.localtime(entry_exit.entry_time).strftime("%H:%M:%S") if entry_exit.entry_time else None,
                "exit_time": timezone.localtime(entry_exit.exit_time).strftime("%H:%M:%S") if entry_exit.exit_time else None,
                "duration": entry_exit.work_duration_formatted if entry_exit.work_duration_seconds else "0ч 0м",
                "is_complete": bool(entry_exit.entry_time and entry_exit.exit_time)
            })
        
        for hikvision_id, entries in entry_exits_by_employee.items():
            employee = Employee.objects.filter(hikvision_id=hikvision_id).first()
            employee_name = employee.name if employee else f"ID_{hikvision_id}"
            result["entry_exits"][hikvision_id] = {
                "employee_name": employee_name,
                "entries_count": len(entries),
                "entries": entries
            }
        
        # 3. Находим проблемы
        for hikvision_id in events_by_employee.keys():
            employee = Employee.objects.filter(hikvision_id=hikvision_id).first()
            employee_name = employee.name if employee else f"ID_{hikvision_id}"
            
            # Проверяем, есть ли полная запись EntryExit
            full_entry_exit = EntryExit.objects.filter(
                hikvision_id=hikvision_id,
                entry_time__gte=start_datetime,
                entry_time__lt=end_datetime,
                exit_time__isnull=False
            ).exists()
            
            if not full_entry_exit:
                events = events_by_employee[hikvision_id]
                result["problems"].append({
                    "hikvision_id": hikvision_id,
                    "employee_name": employee_name,
                    "events_count": len(events),
                    "has_partial_entry_exit": EntryExit.objects.filter(
                        hikvision_id=hikvision_id
                    ).filter(
                        Q(entry_time__gte=start_datetime, entry_time__lt=end_datetime) |
                        Q(exit_time__gte=start_datetime, exit_time__lt=end_datetime)
                    ).exists(),
                    "message": "Есть события CameraEvent, но нет полной записи EntryExit (нет выхода)"
                })
        
        result["summary"] = {
            "total_camera_events": camera_events.count(),
            "total_entry_exits": entry_exits.count(),
            "employees_with_events": len(events_by_employee),
            "employees_with_entry_exits": len(entry_exits_by_employee),
            "problems_count": len(result["problems"])
        }
        
        return Response(result)
    
    @action(detail=False, methods=["get"], url_path="departments-list")
    def departments_list(self, request):
        """
        Возвращает список всех подразделений для выпадающего списка.
        """
        departments = Department.objects.all().order_by('name')
        
        departments_data = []
        for dept in departments:
            full_path = dept.get_full_path()
            # Убираем "АУП" или "АУП > " из начала пути
            if full_path.startswith("АУП > "):
                display_name = full_path[6:]
            elif full_path.startswith("АУП"):
                display_name = full_path[3:].lstrip(" > ")
            else:
                display_name = full_path
            display_name = display_name.lstrip("/ > ")
            
            departments_data.append({
                'id': dept.id,
                'name': display_name,
                'full_path': full_path
            })
        
        return Response(departments_data)
    
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Экспорт записей входов/выходов в Excel с фильтрацией по датам.
        Выводит все даты в диапазоне, пустые ячейки помечаются красным.
        ВСЕГДА использует оптимизированные SQL запросы из sql_reports.py.
        
        Параметры:
        - hikvision_id - фильтр по ID от Hikvision
        - employee_name - фильтр по имени сотрудника (поиск по части имени)
        - department_name - фильтр по подразделению (поиск по части названия)
        - start_date - начальная дата (формат: YYYY-MM-DD или YYYY-MM-DD HH:MM:SS)
        - end_date - конечная дата (формат: YYYY-MM-DD или YYYY-MM-DD HH:MM:SS)
        """
        # Фильтры
        hikvision_id = request.query_params.get("hikvision_id") or request.query_params.get("employee_id")
        employee_name = request.query_params.get("employee_name")
        department_name = request.query_params.get("department_name")
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")
        
        # ВСЕГДА используем оптимизированные SQL запросы
        return self._export_excel_sql_by_department(hikvision_id, employee_name, department_name, start_date_str, end_date_str)
    
    def _export_excel_sql_by_department(self, hikvision_id, employee_name, department_name, start_date_str, end_date_str):
        """
        Оптимизированная версия экспорта с использованием SQL запросов.
        Использует только SQL запросы из sql_reports.py, без ORM.
        Если выбрано подразделение, создает отдельный лист для каждого сотрудника.
        """
        from .sql_reports import generate_comprehensive_attendance_report_sql
        from .utils import get_excluded_hikvision_ids
        from django.db import connection
        from django.db.models import Q
        from datetime import date
        
        # Получаем исключаемые ID через SQL
        excluded_hikvision_ids = get_excluded_hikvision_ids()
        
        # Определяем список сотрудников для экспорта
        employees_to_export = []
        
        if hikvision_id:
            # Если указан конкретный ID, используем его
            employee = Employee.objects.filter(hikvision_id=hikvision_id).first()
            if employee:
                employees_to_export = [employee]
        elif employee_name or department_name:
            # Ищем сотрудников по критериям
            employees_query = Employee.objects.exclude(
                hikvision_id__in=excluded_hikvision_ids
            ).filter(hikvision_id__isnull=False)
            
            if employee_name:
                employees_query = employees_query.filter(name__icontains=employee_name)
            
            if department_name:
                # Ищем по полному пути подразделения или по старому полю
                employees_query = employees_query.filter(
                    Q(department__name__icontains=department_name) |
                    Q(department_old__icontains=department_name)
                )
                # Также проверяем полный путь подразделения через связанные отделы
                department_ids = []
                for dept in Department.objects.filter(name__icontains=department_name):
                    # Получаем все дочерние подразделения
                    def get_all_children(dept_obj):
                        children = [dept_obj.id]
                        for child in dept_obj.children.all():
                            children.extend(get_all_children(child))
                        return children
                    department_ids.extend(get_all_children(dept))
                
                if department_ids:
                    employees_query = employees_query.filter(
                        Q(department_id__in=department_ids) |
                        Q(department__name__icontains=department_name) |
                        Q(department_old__icontains=department_name)
                    )
            
            employees_to_export = list(employees_query.select_related('department').prefetch_related('work_schedules').distinct())
        
        if not employees_to_export:
            # Если не найдено сотрудников, возвращаем пустой Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Нет данных"
            ws.append(["Не найдено сотрудников по указанным критериям"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="no_data.xlsx"'
            return response
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        # Удаляем дефолтный лист
        if wb.worksheets:
            wb.remove(wb.worksheets[0])
        
        # Для каждого сотрудника создаем отдельный лист
        for employee in employees_to_export:
            emp_hikvision_id = employee.hikvision_id
            
            # Получаем данные для этого сотрудника
            results, start_date_obj, end_date_obj = generate_comprehensive_attendance_report_sql(
                hikvision_id=emp_hikvision_id,
                start_date=start_date_str,
                end_date=end_date_str,
                device_name=None,
                excluded_hikvision_ids=excluded_hikvision_ids
            )
            
            # Создаем лист для сотрудника
            # Ограничиваем длину имени листа (Excel ограничение - 31 символ)
            sheet_name = (employee.name or f"ID_{emp_hikvision_id}")[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Вызываем вспомогательную функцию для заполнения листа
            self._fill_employee_sheet(ws, employee, results, start_date_obj, end_date_obj)
        
        # Если нет ни одного листа, создаем пустой
        if len(wb.worksheets) == 0:
            ws = wb.create_sheet(title="Нет данных")
            ws.append(["Не найдено данных"])
        
        # Сохраняем файл
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Генерируем имя файла
        if len(employees_to_export) == 1:
            emp = employees_to_export[0]
            emp_name = re.sub(r'[<>:"/\\|?*]', '_', emp.name or f"ID_{emp.hikvision_id}")
            emp_name = emp_name.replace(' ', '_')
            emp_name = re.sub(r'_+', '_', emp_name)
            if start_date_str and end_date_str:
                start_date_obj = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date_obj = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                start_date_str_formatted = start_date_obj.strftime('%d-%m-%Y')
                end_date_str_formatted = end_date_obj.strftime('%d-%m-%Y')
                filename = f"{emp_name}_с_{start_date_str_formatted}_по_{end_date_str_formatted}.xlsx"
            else:
                filename = f"{emp_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"отчет_по_подразделению_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = FileResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        from urllib.parse import quote
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(filename)}'
        return response
    
    def _fill_employee_sheet(self, ws, employee, results, start_date_obj, end_date_obj):
        """
        Заполняет лист Excel данными для одного сотрудника.
        """
        from datetime import date
        
        # Русские названия дней недели
        WEEKDAYS_RU = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
        
        # Заголовки (новый порядок: ФИО, подразделение, должность, дата, день недели, время графика, тип графика, время входа, время выхода, Продолжительность работы, Опоздание, Ранний уход, Ранний приход, Поздний выход)
        headers = [
            "ФИО",
            "Подразделение",
            "Должность",
            "Дата",
            "День недели",
            "Время графика",
            "Тип графика",
            "Время входа",
            "Время выхода",
            "Продолжительность работы",
            "Опоздание",
            "Ранний уход",
            "Ранний приход",
            "Поздний выход"
        ]
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Получаем информацию о сотруднике
        employee_name = employee.name if employee.name else ""
        department_name = ""
        position = employee.position if employee.position else ""
        schedule = employee.work_schedules.first()
        
        # Получаем название подразделения
        if employee.department:
            full_path = employee.department.get_full_path()
            if full_path.startswith("АУП > "):
                department_name = full_path[6:]
            elif full_path.startswith("АУП"):
                department_name = full_path[3:].lstrip(" > ")
            else:
                department_name = full_path
            department_name = department_name.lstrip("/ > ")
        elif employee.department_old:
            dept_old = employee.department_old
            if dept_old.startswith("АУП/"):
                department_name = dept_old[4:]
            elif dept_old.startswith("АУП"):
                department_name = dept_old[3:].lstrip("/")
            else:
                department_name = dept_old
            department_name = department_name.replace("/", " > ")
            department_name = department_name.lstrip("/ > ")
        
        # Группируем результаты по дате
        data_by_date = {}
        for result in results:
            report_date = result.get('report_date')
            date_key = None
            
            # Обрабатываем разные форматы даты
            if isinstance(report_date, date):
                date_key = report_date
            elif isinstance(report_date, datetime):
                date_key = report_date.date()
            elif isinstance(report_date, str):
                # Пробуем разные форматы
                for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d.%m.%Y"]:
                    try:
                        date_key = datetime.strptime(report_date, fmt).date()
                        break
                    except:
                        continue
            elif report_date is not None:
                # Пробуем преобразовать через str
                try:
                    date_str = str(report_date)
                    for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d.%m.%Y"]:
                        try:
                            date_key = datetime.strptime(date_str, fmt).date()
                            break
                        except:
                            continue
                except:
                    pass
            
            if date_key is None:
                continue
            
            # Если для этой даты уже есть данные, проверяем, нужно ли обновить
            # (берем запись с большей продолжительностью или более позднюю)
            if date_key not in data_by_date:
                data_by_date[date_key] = result
            else:
                # Если есть несколько записей для одной даты, берем ту, у которой больше продолжительность
                existing_duration = data_by_date[date_key].get('total_duration_seconds', 0) or 0
                new_duration = result.get('total_duration_seconds', 0) or 0
                if new_duration > existing_duration:
                    data_by_date[date_key] = result
        
        # Для круглосуточных графиков: если выход на следующий день, создаем запись для следующего дня
        if schedule and schedule.schedule_type == 'round_the_clock':
            additional_data = {}
            for date_key, result_data in list(data_by_date.items()):
                last_exit = result_data.get('last_exit')
                if last_exit:
                    # Преобразуем last_exit в дату
                    exit_date = None
                    if isinstance(last_exit, datetime):
                        exit_date = last_exit.date()
                    elif isinstance(last_exit, str):
                        try:
                            exit_date = datetime.strptime(last_exit, "%Y-%m-%d %H:%M:%S").date()
                        except:
                            try:
                                exit_date = datetime.strptime(last_exit, "%Y-%m-%d").date()
                            except:
                                pass
                    elif hasattr(last_exit, 'date'):
                        exit_date = last_exit.date()
                    
                    # Проверяем, что выход на следующий день и входит в диапазон отчета
                    if exit_date and exit_date > date_key and start_date_obj <= exit_date <= end_date_obj:
                        # Выход на следующий день - создаем запись для следующего дня
                        if exit_date not in data_by_date and exit_date not in additional_data:
                            # Создаем копию результата, но без входа и продолжительности
                            next_day_result = result_data.copy()
                            next_day_result['first_entry'] = None
                            next_day_result['total_duration_seconds'] = 0
                            next_day_result['report_date'] = exit_date
                            next_day_result['late_minutes'] = 0
                            next_day_result['early_leave_minutes'] = 0
                            next_day_result['early_arrival_minutes'] = 0
                            next_day_result['late_departure_minutes'] = 0
                            additional_data[exit_date] = next_day_result
            
            # Добавляем дополнительные записи
            data_by_date.update(additional_data)
        
        # Генерируем все даты в диапазоне
        current_date = start_date_obj
        row_num = 2
        total_duration_hours = 0.0
        total_scheduled_hours = 0.0
        
        main_employee_id = clean_id(employee.hikvision_id) if employee.hikvision_id else None
        
        # Логируем для отладки (можно убрать после проверки)
        logger.debug(f"Обработка данных для сотрудника {employee_name} (ID: {main_employee_id})")
        logger.debug(f"Диапазон дат: {start_date_obj} - {end_date_obj}")
        logger.debug(f"Найдено записей в результатах: {len(results)}")
        logger.debug(f"Уникальные даты в данных: {sorted(data_by_date.keys())}")
        
        while current_date <= end_date_obj:
            date_str = current_date.strftime("%d-%m-%Y")
            weekday_name = WEEKDAYS_RU[current_date.weekday()]
            
            # Получаем информацию о графике для этой даты
            schedule_type_str = ""
            schedule_time_str = ""
            if schedule:
                schedule_type_str = SCHEDULE_TYPE_MAP.get(schedule.schedule_type, "")
                scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, current_date)
                if scheduled_times:
                    scheduled_start, scheduled_end = scheduled_times
                    if schedule.schedule_type == 'round_the_clock':
                        schedule_time_str = "Круглосуточно"
                    else:
                        scheduled_start_local = timezone.localtime(scheduled_start)
                        scheduled_end_local = timezone.localtime(scheduled_end)
                        start_str = scheduled_start_local.strftime('%H:%M')
                        end_str = scheduled_end_local.strftime('%H:%M')
                        schedule_time_str = f"{start_str}-{end_str}"
            
            # Ищем данные для текущей даты
            found_data = data_by_date.get(current_date)
            
            if found_data:
                # Есть данные для этой даты
                first_entry = found_data.get('first_entry')
                last_exit = found_data.get('last_exit')
                total_duration_seconds = found_data.get('total_duration_seconds', 0) or 0
                late_minutes = found_data.get('late_minutes', 0) or 0
                early_leave_minutes = found_data.get('early_leave_minutes', 0) or 0
                early_arrival_minutes = found_data.get('early_arrival_minutes', 0) or 0
                late_departure_minutes = found_data.get('late_departure_minutes', 0) or 0
                
                # Форматируем время входа и выхода
                entry_time_str = ""
                exit_time_str = ""
                
                if first_entry:
                    if isinstance(first_entry, datetime):
                        if timezone.is_naive(first_entry):
                            if ALMATY_TZ:
                                entry_time_aware = first_entry.replace(tzinfo=ALMATY_TZ)
                            else:
                                entry_time_aware = timezone.make_aware(first_entry)
                            entry_time_local = timezone.localtime(entry_time_aware)
                        else:
                            entry_time_local = timezone.localtime(first_entry)
                        # Для круглосуточных графиков: вход всегда только время
                        if schedule and schedule.schedule_type == 'round_the_clock':
                            entry_time_str = entry_time_local.strftime("%H:%M:%S")
                        else:
                            entry_time_str = entry_time_local.strftime("%H:%M:%S")
                    else:
                        entry_time_str = str(first_entry)
                
                if last_exit:
                    if isinstance(last_exit, datetime):
                        if timezone.is_naive(last_exit):
                            if ALMATY_TZ:
                                exit_time_aware = last_exit.replace(tzinfo=ALMATY_TZ)
                            else:
                                exit_time_aware = timezone.make_aware(last_exit)
                            exit_time_local = timezone.localtime(exit_time_aware)
                        else:
                            exit_time_local = timezone.localtime(last_exit)
                        # Для круглосуточных графиков: если выход на следующий день, показываем с датой
                        if schedule and schedule.schedule_type == 'round_the_clock':
                            exit_date = exit_time_local.date()
                            if exit_date > current_date:
                                # Выход на следующий день - показываем с полной датой
                                exit_time_str = exit_time_local.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                # Выход в тот же день - только время
                                exit_time_str = exit_time_local.strftime("%H:%M:%S")
                        else:
                            exit_time_str = exit_time_local.strftime("%H:%M:%S")
                    else:
                        exit_time_str = str(last_exit)
                
                # Продолжительность
                duration_hours = int(total_duration_seconds) // 3600
                duration_minutes = (int(total_duration_seconds) % 3600) // 60
                duration_str = f"{duration_hours}ч {duration_minutes}м" if total_duration_seconds > 0 else ""
                duration_hours_float = duration_hours + (duration_minutes / 60.0)
                total_duration_hours += duration_hours_float
                
                # Форматируем опоздание и ранний уход
                late_str = ""
                if late_minutes and late_minutes > 0:
                    late_hours = int(late_minutes) // 60
                    late_mins = int(late_minutes) % 60
                    if late_hours > 0:
                        late_str = f"{late_hours}ч {late_mins}м"
                    else:
                        late_str = f"{late_mins}м"
                
                early_leave_str = ""
                if early_leave_minutes and early_leave_minutes > 0:
                    early_hours = int(early_leave_minutes) // 60
                    early_mins = int(early_leave_minutes) % 60
                    if early_hours > 0:
                        early_leave_str = f"{early_hours}ч {early_mins}м"
                    else:
                        early_leave_str = f"{early_mins}м"
                
                # Форматируем ранний приход и поздний выход
                early_arrival_str = ""
                if early_arrival_minutes and early_arrival_minutes > 0:
                    early_arr_hours = int(early_arrival_minutes) // 60
                    early_arr_mins = int(early_arrival_minutes) % 60
                    if early_arr_hours > 0:
                        early_arrival_str = f"{early_arr_hours}ч {early_arr_mins}м"
                    else:
                        early_arrival_str = f"{early_arr_mins}м"
                
                late_departure_str = ""
                if late_departure_minutes and late_departure_minutes > 0:
                    late_dep_hours = int(late_departure_minutes) // 60
                    late_dep_mins = int(late_departure_minutes) % 60
                    if late_dep_hours > 0:
                        late_departure_str = f"{late_dep_hours}ч {late_dep_mins}м"
                    else:
                        late_departure_str = f"{late_dep_mins}м"
                
                # Суммируем время по графику для этого дня
                if schedule:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, current_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        scheduled_duration = (scheduled_end - scheduled_start).total_seconds() / 3600.0
                        total_scheduled_hours += scheduled_duration
                
                # Новый порядок колонок: ФИО, подразделение, должность, дата, день недели, время графика, тип графика, время входа, время выхода, Продолжительность работы, Опоздание, Ранний уход, Ранний приход, Поздний выход
                ws.append([
                    employee_name,
                    department_name,
                    position,
                    date_str,
                    weekday_name,
                    schedule_time_str,
                    schedule_type_str,
                    entry_time_str,
                    exit_time_str,
                    duration_str,
                    late_str,
                    early_leave_str,
                    early_arrival_str,
                    late_departure_str,
                ])
                
                # Применяем стили (новый порядок колонок: A-N)
                for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N'], 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # Красный цвет для пустых ячеек времени входа/выхода и при малой продолжительности
                    if col_letter == 'H':  # Время входа (колонка H)
                        if not entry_time_str or (duration_hours_float > 0 and duration_hours_float < 2.0):
                            cell.fill = red_fill
                    elif col_letter == 'I':  # Время выхода (колонка I)
                        if not exit_time_str or (duration_hours_float > 0 and duration_hours_float < 2.0):
                            cell.fill = red_fill
                    elif col_letter == 'J':  # Продолжительность работы (колонка J)
                        if duration_hours_float > 0 and duration_hours_float < 2.0:
                            cell.fill = red_fill
                        elif not duration_str:
                            cell.fill = red_fill
                    elif col_letter == 'K':  # Опоздание (колонка K)
                        if late_str:  # Если есть опоздание - красный цвет
                            cell.fill = red_fill
                    elif col_letter == 'L':  # Ранний уход (колонка L)
                        if early_leave_str:  # Если есть ранний уход - красный цвет
                            cell.fill = red_fill
                    elif col_letter == 'M':  # Ранний приход (колонка M)
                        if early_arrival_str:  # Если есть ранний приход - зеленый цвет
                            cell.fill = green_fill
                    elif col_letter == 'N':  # Поздний выход (колонка N)
                        if late_departure_str:  # Если есть поздний выход - зеленый цвет
                            cell.fill = green_fill
            else:
                # Нет данных для этой даты - создаем пустую строку с красным
                # Новый порядок колонок: ФИО, подразделение, должность, дата, день недели, время графика, тип графика, время входа, время выхода, Продолжительность работы, Опоздание, Ранний уход, Ранний приход, Поздний выход
                ws.append([
                    employee_name,
                    department_name,
                    position,
                    date_str,
                    weekday_name,
                    schedule_time_str,
                    schedule_type_str,
                    "",  # Время входа - пустое, будет красным
                    "",  # Время выхода - пустое, будет красным
                    "",  # Продолжительность - пустая, будет красной
                    "",  # Опоздание - пустое
                    "",  # Ранний уход - пустое
                    "",  # Ранний приход - пустое
                    "",  # Поздний выход - пустое
                ])
                
                # Применяем стили и красный цвет для пустых ячеек (новый порядок колонок: A-N)
                for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N'], 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # Красный цвет для пустых ячеек времени входа/выхода и продолжительности
                    if col_letter == 'H':  # Время входа - всегда красное
                        cell.fill = red_fill
                    elif col_letter == 'I':  # Время выхода - всегда красное
                        cell.fill = red_fill
                    elif col_letter == 'J':  # Продолжительность - всегда красная для пустых строк
                        cell.fill = red_fill
                
                # Для пустых дней также проверяем график
                if schedule:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, current_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        scheduled_duration = (scheduled_end - scheduled_start).total_seconds() / 3600.0
                        total_scheduled_hours += scheduled_duration
            
            row_num += 1
            current_date += timedelta(days=1)
        
        # Добавляем итоговую строку (используем ту же логику, что была раньше)
        total_hours = int(total_duration_hours)
        total_minutes = int((total_duration_hours - total_hours) * 60)
        total_duration_str = f"{total_hours}ч {total_minutes}м" if total_duration_hours > 0 else ""
        
        # Пересчитываем общее время работы по графику
        recalculated_scheduled_hours = 0.0
        schedule_time_display = ""
        
        if schedule:
            if schedule.schedule_type == 'regular' and schedule.start_time and schedule.end_time:
                start_str = schedule.start_time.strftime('%H:%M')
                end_str = schedule.end_time.strftime('%H:%M')
                schedule_time_display = f"{start_str}-{end_str}"
                
                start_datetime = datetime.combine(start_date_obj, schedule.start_time)
                end_datetime = datetime.combine(start_date_obj, schedule.end_time)
                if schedule.end_time < schedule.start_time:
                    end_datetime += timedelta(days=1)
                shift_duration_hours = (end_datetime - start_datetime).total_seconds() / 3600.0
                
                working_days_count = 0
                check_date = start_date_obj
                while check_date <= end_date_obj:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, check_date)
                    if scheduled_times:
                        working_days_count += 1
                    check_date += timedelta(days=1)
                
                recalculated_scheduled_hours = shift_duration_hours * working_days_count
            elif schedule.schedule_type == 'floating':
                check_date = start_date_obj
                while check_date <= end_date_obj:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, check_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        shift_duration = (scheduled_end - scheduled_start).total_seconds() / 3600.0
                        recalculated_scheduled_hours += shift_duration
                        
                        if not schedule_time_display:
                            scheduled_start_local = timezone.localtime(scheduled_start)
                            scheduled_end_local = timezone.localtime(scheduled_end)
                            start_str = scheduled_start_local.strftime('%H:%M')
                            end_str = scheduled_end_local.strftime('%H:%M')
                            schedule_time_display = f"{start_str}-{end_str}"
                    check_date += timedelta(days=1)
            elif schedule.schedule_type == 'round_the_clock':
                # Для круглосуточных графиков используем правильную формулу расчета:
                # Каждая рабочая смена = 24 часа (сутки)
                # Считаем количество дней с запланированными сменами
                
                # Определяем время начала смены (по умолчанию 09:00, но может быть другое)
                shift_start_time = schedule.start_time if schedule.start_time else time(9, 0)
                
                # Для круглосуточных графиков считаем уникальные смены
                # Смена определяется по дате начала (дню, когда она начинается)
                shifts_dates = set()  # Множество для хранения дат начала смен
                
                check_date = start_date_obj
                while check_date <= end_date_obj:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, check_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        # Дата начала смены определяет уникальную смену
                        shift_start_date = scheduled_start.date()
                        shifts_dates.add(shift_start_date)
                    check_date += timedelta(days=1)
                
                # Количество уникальных смен (каждая смена = 24 часа)
                # Формула: количество смен × 24 часа
                number_of_shifts = len(shifts_dates)
                recalculated_scheduled_hours = 24.0 * number_of_shifts
                schedule_time_display = "Круглосуточно"
        
        final_scheduled_hours = recalculated_scheduled_hours if recalculated_scheduled_hours > 0 else total_scheduled_hours
        
        if not schedule_time_display and schedule:
            if schedule.start_time and schedule.end_time:
                start_str = schedule.start_time.strftime('%H:%M')
                end_str = schedule.end_time.strftime('%H:%M')
                schedule_time_display = f"{start_str}-{end_str}"
            elif schedule.schedule_type == 'round_the_clock':
                schedule_time_display = "Круглосуточно"
        
        if final_scheduled_hours > 0:
            scheduled_hours = int(final_scheduled_hours)
            scheduled_minutes = int(round((final_scheduled_hours - scheduled_hours) * 60))
            scheduled_duration_str = f"{scheduled_hours}ч" if scheduled_minutes == 0 else f"{scheduled_hours}ч {scheduled_minutes}м"
        else:
            scheduled_duration_str = ""
        
        schedule_type_value = ""
        if scheduled_duration_str:
            schedule_type_value = f"должен отработать: {scheduled_duration_str}"
        
        # Добавляем итоговую строку (новый порядок колонок: ФИО, подразделение, должность, дата, день недели, время графика, тип графика, время входа, время выхода, Продолжительность работы)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_font = Font(bold=True, size=12)
        
        ws.append([
            "ИТОГО:",
            "",
            "",
            "",
            "",
            schedule_time_display if schedule_time_display else "",
            schedule_type_value,
            "",
            "",
            total_duration_str,
        ])
        
        total_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Автоподбор ширины колонок (новый порядок: ФИО, подразделение, должность, дата, день недели, время графика, тип графика, время входа, время выхода, Продолжительность работы)
        column_widths = {
            "A": 30,  # ФИО
            "B": 35,  # Подразделение
            "C": 20,  # Должность
            "D": 15,  # Дата
            "E": 15,  # День недели
            "F": 20,  # Время графика
            "G": 25,  # Тип графика
            "H": 20,  # Время входа
            "I": 20,  # Время выхода
            "J": 25   # Продолжительность работы
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        ws.row_dimensions[1].height = 30
    
    @action(detail=False, methods=["post"], url_path="full-recalculate")
    def full_recalculate(self, request):
        """
        Полный пересчет всех входов и выходов с 1 декабря и пересчет статистики.
        Запускает пересчет EntryExit из CameraEvent, затем пересчет статистики посещаемости.
        
        Параметры (опционально, в теле запроса или query params):
        - start_date: Дата начала пересчета в формате YYYY-MM-DD (по умолчанию: 1 декабря текущего года)
        - end_date: Дата окончания пересчета в формате YYYY-MM-DD (по умолчанию: сегодня)
        """
        try:
            # Получаем параметры
            if hasattr(request, 'data') and request.data:
                start_date_str = request.data.get('start_date')
                end_date_str = request.data.get('end_date')
            else:
                start_date_str = request.query_params.get('start_date')
                end_date_str = request.query_params.get('end_date')
            
            # Парсим даты
            start_date = None
            end_date = None
            
            if start_date_str:
                try:
                    if ' ' in start_date_str or 'T' in start_date_str:
                        start_date = datetime.strptime(start_date_str.replace('T', ' '), "%Y-%m-%d %H:%M:%S")
                    else:
                        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                except ValueError as e:
                    return JsonResponse({
                        "status": "error",
                        "message": f"Неверный формат start_date: {e}"
                    }, status=400)
            
            if end_date_str:
                try:
                    if ' ' in end_date_str or 'T' in end_date_str:
                        end_date = datetime.strptime(end_date_str.replace('T', ' '), "%Y-%m-%d %H:%M:%S")
                    else:
                        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                except ValueError as e:
                    return JsonResponse({
                        "status": "error",
                        "message": f"Неверный формат end_date: {e}"
                    }, status=400)
            
            # Если даты не указаны, используем значения по умолчанию: с 1 декабря по сегодня
            if not start_date:
                today = timezone.now().date()
                start_date = datetime.combine(datetime(today.year, 12, 1).date(), datetime.min.time())
                start_date = timezone.make_aware(start_date)
                logger.info(f"Using default start_date: {start_date}")
            
            if not end_date:
                end_date = timezone.now()
                logger.info(f"Using default end_date: {end_date}")
            
            # Делаем даты aware, если нужно
            if start_date and timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date)
            if end_date and timezone.is_naive(end_date):
                end_date = timezone.make_aware(end_date)
            
            # Запускаем пересчет в отдельном потоке, чтобы не блокировать ответ
            import threading
            
            def run_full_recalculate():
                try:
                    logger.info(f"Начинаем полный пересчет с {start_date} по {end_date}")
                    
                    # Шаг 1: Пересчитываем EntryExit из CameraEvent
                    logger.info("Шаг 1: Пересчет EntryExit из CameraEvent...")
                    result_entries = recalculate_entries_exits(start_date=start_date, end_date=end_date)
                    logger.info(f"EntryExit пересчет завершен: создано={result_entries.get('created', 0)}, обновлено={result_entries.get('updated', 0)}")
                    
                    # Шаг 2: Пересчитываем статистику посещаемости
                    logger.info("Шаг 2: Пересчет статистики посещаемости...")
                    import sys
                    from pathlib import Path
                    
                    project_root = Path(__file__).resolve().parent.parent
                    if str(project_root) not in sys.path:
                        sys.path.insert(0, str(project_root))
                    
                    from recalculate_attendance_stats import recalculate_attendance_stats
                    
                    # Используем только дату (без времени) для пересчета статистики
                    start_date_only = start_date.date() if hasattr(start_date, 'date') else start_date
                    recalculate_attendance_stats(start_date=start_date_only)
                    
                    logger.info("Полный пересчет завершен успешно!")
                except Exception as e:
                    logger.error(f"Ошибка при полном пересчете: {e}", exc_info=True)
            
            thread = threading.Thread(target=run_full_recalculate, daemon=True)
            thread.start()
            
            return JsonResponse({
                "status": "success",
                "message": "Полный пересчет запущен в фоновом режиме",
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
                "note": "Пересчет включает: 1) EntryExit из CameraEvent, 2) Статистику посещаемости"
            })
        except Exception as e:
            logger.error(f"Ошибка при запуске полного пересчета: {e}", exc_info=True)
            return JsonResponse({
                "status": "error",
                "message": str(e)
            }, status=500)


# DepartmentViewSet вынесен в viewsets/department.py
# Импортируется выше


# DepartmentViewSet вынесен в viewsets/department.py
# Импортируется выше

class AttendanceStatsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet для экспорта статистики посещаемости по подразделениям.
    """
    queryset = Employee.objects.none()  # Не используется для стандартных операций
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Экспорт статистики посещаемости по подразделениям в Excel.
        Фильтрует сотрудников по department_id (может быть несколько) и экспортирует их данные.
        
        Параметры:
        - department_id - ID подразделения (можно указать несколько раз)
        - start_date - начальная дата (формат: YYYY-MM-DD)
        - end_date - конечная дата (формат: YYYY-MM-DD)
        """
        # Получаем параметры
        department_ids = request.query_params.getlist("department_id")
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")
        
        # Валидация
        if not department_ids:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ошибка"
            ws.append(["Не указаны ID подразделений. Используйте параметр department_id."])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="error.xlsx"'
            return response
        
        return self._export_excel_by_department_ids(department_ids, start_date_str, end_date_str)
    
    def _export_excel_by_department_ids(self, department_ids, start_date_str, end_date_str):
        """
        Экспорт данных по ID подразделений с использованием SQL запросов.
        """
        from .sql_reports import generate_comprehensive_attendance_report_sql
        from .utils import get_excluded_hikvision_ids
        from django.db.models import Q
        
        # Получаем исключаемые ID
        excluded_hikvision_ids = get_excluded_hikvision_ids()
        
        # Функция для получения всех дочерних подразделений
        def get_all_children(dept_obj):
            children = [dept_obj.id]
            for child in dept_obj.children.all():
                children.extend(get_all_children(child))
            return children
        
        # Собираем все ID подразделений (включая дочерние)
        all_department_ids = []
        for dept_id in department_ids:
            try:
                dept_id_int = int(dept_id)
                dept = Department.objects.filter(id=dept_id_int).first()
                if dept:
                    all_department_ids.extend(get_all_children(dept))
            except (ValueError, TypeError):
                continue
        
        # Убираем дубликаты
        all_department_ids = list(set(all_department_ids))
        
        if not all_department_ids:
            # Если не найдено подразделений, возвращаем пустой Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Нет данных"
            ws.append(["Не найдено подразделений по указанным ID"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="no_data.xlsx"'
            return response
        
        # Ищем сотрудников по подразделениям
        employees_query = Employee.objects.exclude(
            hikvision_id__in=excluded_hikvision_ids
        ).filter(
            hikvision_id__isnull=False,
            department_id__in=all_department_ids
        )
        
        employees_to_export = list(
            employees_query.select_related('department').prefetch_related('work_schedules').distinct()
        )
        
        if not employees_to_export:
            # Если не найдено сотрудников, возвращаем пустой Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Нет данных"
            ws.append(["Не найдено сотрудников в указанных подразделениях"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="no_data.xlsx"'
            return response
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        # Удаляем дефолтный лист
        if wb.worksheets:
            wb.remove(wb.worksheets[0])
        
        # Для каждого сотрудника создаем отдельный лист
        for employee in employees_to_export:
            emp_hikvision_id = employee.hikvision_id
            
            # Получаем данные для этого сотрудника
            results, start_date_obj, end_date_obj = generate_comprehensive_attendance_report_sql(
                hikvision_id=emp_hikvision_id,
                start_date=start_date_str,
                end_date=end_date_str,
                device_name=None,
                excluded_hikvision_ids=excluded_hikvision_ids
            )
            
            # Создаем лист для сотрудника
            # Ограничиваем длину имени листа (Excel ограничение - 31 символ)
            sheet_name = (employee.name or f"ID_{emp_hikvision_id}")[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Используем метод из EntryExitViewSet для заполнения листа
            # Создаем временный экземпляр для вызова метода
            entry_exit_viewset = EntryExitViewSet()
            entry_exit_viewset._fill_employee_sheet(ws, employee, results, start_date_obj, end_date_obj)
        
        # Если нет ни одного листа, создаем пустой
        if len(wb.worksheets) == 0:
            ws = wb.create_sheet(title="Нет данных")
            ws.append(["Не найдено данных"])
        
        # Сохраняем файл
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Генерируем имя файла
        if len(employees_to_export) == 1:
            emp = employees_to_export[0]
            emp_name = re.sub(r'[<>:"/\\|?*]', '_', emp.name or f"ID_{emp.hikvision_id}")
            emp_name = emp_name.replace(' ', '_')
            emp_name = re.sub(r'_+', '_', emp_name)
            if start_date_str and end_date_str:
                start_date_obj = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date_obj = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                start_date_str_formatted = start_date_obj.strftime('%d-%m-%Y')
                end_date_str_formatted = end_date_obj.strftime('%d-%m-%Y')
                filename = f"{emp_name}_с_{start_date_str_formatted}_по_{end_date_str_formatted}.xlsx"
            else:
                filename = f"{emp_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"отчет_по_подразделениям_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = FileResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        from urllib.parse import quote
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(filename)}'
        return response


# DepartmentViewSet вынесен в viewsets/department.py
# Импортируется выше
        header_font = Font(bold=True, color="FFFFFF", size=12)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Получаем информацию о сотруднике и его графике для всех строк
        employee = None
        schedule = None
        employee_name = ""
        department_name = ""
        position = ""
        
        if main_employee_id:
            try:
                # Используем clean_id для нормализации ID
                clean_emp_id = clean_id(main_employee_id)
                employee = Employee.objects.filter(hikvision_id=clean_emp_id).first()
                if not employee:
                    # Пробуем найти без clean_id (на случай если в БД хранится с ведущими нулями)
                    employee = Employee.objects.filter(hikvision_id=main_employee_id).first()
                if employee:
                    schedule = employee.work_schedules.first()
                    employee_name = employee.name if employee.name else ""
                    
                    # Получаем название подразделения
                    if employee.department:
                        full_path = employee.department.get_full_path()
                        # Убираем "АУП" или "АУП > " из начала пути
                        if full_path.startswith("АУП > "):
                            department_name = full_path[6:]  # Убираем "АУП > "
                        elif full_path.startswith("АУП"):
                            department_name = full_path[3:].lstrip(" > ")  # Убираем "АУП" и возможные разделители
                        else:
                            department_name = full_path
                        # Убираем ведущие разделители на случай, если они остались
                        department_name = department_name.lstrip("/ > ")
                    elif employee.department_old:
                        dept_old = employee.department_old
                        # Убираем "АУП/" из начала для старого поля и заменяем "/" на " > "
                        if dept_old.startswith("АУП/"):
                            department_name = dept_old[4:]  # Убираем "АУП/"
                        elif dept_old.startswith("АУП"):
                            department_name = dept_old[3:].lstrip("/")  # Убираем "АУП" и возможные разделители
                        else:
                            department_name = dept_old
                        # Заменяем "/" на " > " и убираем ведущие разделители
                        department_name = department_name.replace("/", " > ")
                        department_name = department_name.lstrip("/ > ")
                    
                    # Получаем должность
                    position = employee.position if employee.position else ""
            except Exception as e:
                logger.warning(f"Ошибка при получении информации о сотруднике: {e}")
                pass
        
        # Генерируем все даты в диапазоне
        current_date = start_date_obj
        row_num = 2
        total_duration_hours = 0.0  # Общее время работы для итоговой строки
        total_scheduled_hours = 0.0  # Общее время по графику (должен отработать)
        
        while current_date <= end_date_obj:
            date_str = current_date.strftime("%d-%m-%Y")
            
            # Ищем данные для текущей даты
            found_data = None
            if main_employee_id:
                key = (main_employee_id, current_date)
                found_data = data_by_employee_date.get(key)
            
            if found_data:
                # Есть данные для этой даты
                first_entry = found_data.get('first_entry')
                last_exit = found_data.get('last_exit')
                total_duration_seconds = found_data.get('total_duration_seconds', 0) or 0
                
                # Форматируем время входа и выхода
                # Результаты SQL запроса возвращают время в местном часовом поясе (Asia/Almaty)
                # но как naive datetime, поэтому нужно правильно интерпретировать его
                entry_time_str = ""
                exit_time_str = ""
                if first_entry:
                    if isinstance(first_entry, datetime):
                        # Если время naive, интерпретируем его как местное время (Asia/Almaty)
                        if timezone.is_naive(first_entry):
                            # Создаем aware datetime в местном часовом поясе
                            if ALMATY_TZ:
                                entry_time_aware = first_entry.replace(tzinfo=ALMATY_TZ)
                            else:
                                # Используем настройки Django (TIME_ZONE = 'Asia/Almaty')
                                entry_time_aware = timezone.make_aware(first_entry)
                            entry_time_local = timezone.localtime(entry_time_aware)
                        else:
                            entry_time_local = timezone.localtime(first_entry)
                        # Для круглосуточных графиков: вход всегда только время
                        if schedule and schedule.schedule_type == 'round_the_clock':
                            entry_time_str = entry_time_local.strftime("%H:%M:%S")
                        else:
                            entry_time_str = entry_time_local.strftime("%H:%M:%S")
                    else:
                        entry_time_str = str(first_entry)
                
                if last_exit:
                    if isinstance(last_exit, datetime):
                        # Если время naive, интерпретируем его как местное время (Asia/Almaty)
                        if timezone.is_naive(last_exit):
                            # Создаем aware datetime в местном часовом поясе
                            if ALMATY_TZ:
                                exit_time_aware = last_exit.replace(tzinfo=ALMATY_TZ)
                            else:
                                # Используем настройки Django (TIME_ZONE = 'Asia/Almaty')
                                exit_time_aware = timezone.make_aware(last_exit)
                            exit_time_local = timezone.localtime(exit_time_aware)
                        else:
                            exit_time_local = timezone.localtime(last_exit)
                        # Для круглосуточных графиков: если выход на следующий день, показываем с датой
                        if schedule and schedule.schedule_type == 'round_the_clock':
                            exit_date = exit_time_local.date()
                            if exit_date > current_date:
                                # Выход на следующий день - показываем с полной датой
                                exit_time_str = exit_time_local.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                # Выход в тот же день - только время
                                exit_time_str = exit_time_local.strftime("%H:%M:%S")
                        else:
                            exit_time_str = exit_time_local.strftime("%H:%M:%S")
                    else:
                        exit_time_str = str(last_exit)
                
                # Продолжительность
                duration_hours = int(total_duration_seconds) // 3600
                duration_minutes = (int(total_duration_seconds) % 3600) // 60
                duration_str = f"{duration_hours}ч {duration_minutes}м" if total_duration_seconds > 0 else ""
                
                # Вычисляем продолжительность в часах для проверки
                duration_hours_float = duration_hours + (duration_minutes / 60.0)
                
                # Суммируем общее время работы
                total_duration_hours += duration_hours_float
                
                # Суммируем время по графику для этого дня
                if schedule:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, current_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        scheduled_duration = (scheduled_end - scheduled_start).total_seconds() / 3600.0
                        total_scheduled_hours += scheduled_duration
                
                ws.append([
                    date_str,
                    employee_name,
                    department_name,
                    position,
                    "",  # Тип графика (для обычных строк пусто)
                    "",  # Время графика (для обычных строк пусто)
                    entry_time_str,
                    exit_time_str,
                    duration_str,
                    "",  # Устройство входа (не в SQL запросе)
                    "",  # Устройство выхода (не в SQL запросе)
                ])
                
                # Применяем стили и красный цвет для малой продолжительности
                for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # Красный цвет для пустых ячеек времени входа/выхода и при малой продолжительности
                    if col_letter == 'G':  # Время входа
                        if not entry_time_str or (duration_hours_float > 0 and duration_hours_float < 2.0):
                            cell.fill = red_fill
                    elif col_letter == 'H':  # Время выхода
                        if not exit_time_str or (duration_hours_float > 0 and duration_hours_float < 2.0):
                            cell.fill = red_fill
                    elif col_letter == 'I':  # Продолжительность работы
                        if duration_hours_float > 0 and duration_hours_float < 2.0:
                            cell.fill = red_fill
                        elif not duration_str:
                            cell.fill = red_fill
            else:
                # Нет данных для этой даты - создаем пустую строку с красным
                ws.append([
                    date_str,
                    employee_name,
                    department_name,
                    position,
                    "",  # Тип графика
                    "",  # Время графика
                    "",  # Время входа - пустое, будет красным
                    "",  # Время выхода - пустое, будет красным
                    "",
                    "",
                    "",
                ])
                
                # Применяем стили и красный цвет для пустых ячеек
                for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # Красный цвет для пустых ячеек времени входа/выхода и продолжительности
                    if col_letter == 'G':  # Время входа - всегда красное
                        cell.fill = red_fill
                    elif col_letter == 'H':  # Время выхода - всегда красное
                        cell.fill = red_fill
                    elif col_letter == 'I':  # Продолжительность - всегда красная для пустых строк
                        cell.fill = red_fill
                
                # Для пустых дней также проверяем график
                if schedule:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, current_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        scheduled_duration = (scheduled_end - scheduled_start).total_seconds() / 3600.0
                        total_scheduled_hours += scheduled_duration
            
            row_num += 1
            current_date += timedelta(days=1)
        
        # Форматируем общее время работы
        total_hours = int(total_duration_hours)
        total_minutes = int((total_duration_hours - total_hours) * 60)
        total_duration_str = f"{total_hours}ч {total_minutes}м" if total_duration_hours > 0 else ""
        
        # Пересчитываем общее время работы по графику для всего периода
        recalculated_scheduled_hours = 0.0
        schedule_time_display = ""
        
        if schedule:
            # Определяем формат времени графика
            if schedule.schedule_type == 'regular' and schedule.start_time and schedule.end_time:
                start_str = schedule.start_time.strftime('%H:%M')
                end_str = schedule.end_time.strftime('%H:%M')
                schedule_time_display = f"{start_str}-{end_str}"
                
                # Рассчитываем продолжительность одной смены
                start_datetime = datetime.combine(start_date_obj, schedule.start_time)
                end_datetime = datetime.combine(start_date_obj, schedule.end_time)
                if schedule.end_time < schedule.start_time:
                    # Ночная смена - добавляем день
                    end_datetime += timedelta(days=1)
                shift_duration_hours = (end_datetime - start_datetime).total_seconds() / 3600.0
                
                # Подсчитываем количество рабочих дней в периоде
                working_days_count = 0
                check_date = start_date_obj
                while check_date <= end_date_obj:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, check_date)
                    if scheduled_times:
                        working_days_count += 1
                    check_date += timedelta(days=1)
                
                recalculated_scheduled_hours = shift_duration_hours * working_days_count
                
            elif schedule.schedule_type == 'floating':
                # Для плавающего графика суммируем время по каждой смене
                check_date = start_date_obj
                while check_date <= end_date_obj:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, check_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        shift_duration = (scheduled_end - scheduled_start).total_seconds() / 3600.0
                        recalculated_scheduled_hours += shift_duration
                        
                        # Формируем строку времени графика из первой найденной смены
                        if not schedule_time_display:
                            scheduled_start_local = timezone.localtime(scheduled_start)
                            scheduled_end_local = timezone.localtime(scheduled_end)
                            start_str = scheduled_start_local.strftime('%H:%M')
                            end_str = scheduled_end_local.strftime('%H:%M')
                            schedule_time_display = f"{start_str}-{end_str}"
                    check_date += timedelta(days=1)
                
            elif schedule.schedule_type == 'round_the_clock':
                # Для круглосуточных графиков используем правильную формулу расчета:
                # Каждая рабочая смена = 24 часа (сутки)
                # Считаем количество уникальных смен (по дате начала смены)
                
                # Определяем время начала смены (по умолчанию 09:00, но может быть другое)
                shift_start_time = schedule.start_time if schedule.start_time else time(9, 0)
                
                # Для круглосуточных графиков считаем уникальные смены
                # Смена определяется по дате начала (дню, когда она начинается)
                shifts_dates = set()  # Множество для хранения дат начала смен
                
                check_date = start_date_obj
                while check_date <= end_date_obj:
                    scheduled_times = ScheduleMatcher.get_scheduled_time_for_date(schedule, check_date)
                    if scheduled_times:
                        scheduled_start, scheduled_end = scheduled_times
                        # Дата начала смены определяет уникальную смену
                        shift_start_date = scheduled_start.date()
                        shifts_dates.add(shift_start_date)
                    check_date += timedelta(days=1)
                
                # Количество уникальных смен (каждая смена = 24 часа)
                # Формула: количество смен × 24 часа
                number_of_shifts = len(shifts_dates)
                recalculated_scheduled_hours = 24.0 * number_of_shifts
                schedule_time_display = "Круглосуточно"
        
        # Используем пересчитанное время, если оно есть и больше 0
        if recalculated_scheduled_hours > 0:
            final_scheduled_hours = recalculated_scheduled_hours
        else:
            # Используем сумму из цикла
            final_scheduled_hours = total_scheduled_hours
        
        # Если schedule_time_display пустой, но есть график, формируем его из графика
        if not schedule_time_display and schedule:
            if schedule.start_time and schedule.end_time:
                start_str = schedule.start_time.strftime('%H:%M')
                end_str = schedule.end_time.strftime('%H:%M')
                schedule_time_display = f"{start_str}-{end_str}"
            elif schedule.schedule_type == 'round_the_clock':
                schedule_time_display = "Круглосуточно"
        
        # Форматируем время по графику (должен отработать)
        if final_scheduled_hours > 0:
            scheduled_hours = int(final_scheduled_hours)
            scheduled_minutes = int(round((final_scheduled_hours - scheduled_hours) * 60))
            scheduled_duration_str = f"{scheduled_hours}ч" if scheduled_minutes == 0 else f"{scheduled_hours}ч {scheduled_minutes}м"
        else:
            scheduled_duration_str = ""
        
        # Формируем значения для итоговой строки
        # В колонке "Тип графика" (колонка 5) - "должен отработать: Xч"
        # В колонке "Время графика" (колонка 6) - время графика (например "09:00-18:00")
        schedule_type_value = ""
        if scheduled_duration_str:
            schedule_type_value = f"должен отработать: {scheduled_duration_str}"
        
        # Стили для итоговой строки
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_font = Font(bold=True, size=12)
        
        # Добавляем итоговую строку
        ws.append([
            "ИТОГО:",  # Колонка A - Дата
            "",  # Колонка B - ФИО
            "",  # Колонка C - Подразделение
            "",  # Колонка D - Должность
            schedule_type_value,  # Колонка E - Тип графика ("должен отработать: Xч")
            schedule_time_display if schedule_time_display else "",  # Колонка F - Время графика
            "",  # Колонка G - Время входа
            "",  # Колонка H - Время выхода
            total_duration_str,  # Колонка I - Продолжительность работы
            "",  # Колонка J - Устройство входа
            "",  # Колонка K - Устройство выхода
        ])
        
        # Применяем стили к итоговой строке
        total_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Автоподбор ширины колонок
        column_widths = {
            "A": 15,  # Дата
            "B": 30,  # ФИО
            "C": 35,  # Подразделение
            "D": 20,  # Должность
            "E": 25,  # Тип графика
            "F": 20,  # Время графика
            "G": 20,  # Время входа
            "H": 20,  # Время выхода
            "I": 25,  # Продолжительность работы
            "J": 20,  # Устройство входа
            "K": 20,  # Устройство выхода
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Высота строки заголовка
        ws.row_dimensions[1].height = 30
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Генерируем имя файла с датой
        filename = f"entries_exits_sql_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Возвращаем файл
        response = FileResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @action(detail=False, methods=["post"], url_path="full-recalculate")
    def full_recalculate(self, request):
        """
        Полный пересчет всех входов и выходов с 1 декабря и пересчет статистики.
        Запускает пересчет EntryExit из CameraEvent, затем пересчет статистики посещаемости.
        
        Параметры (опционально, в теле запроса или query params):
        - start_date: Дата начала пересчета в формате YYYY-MM-DD (по умолчанию: 1 декабря текущего года)
        - end_date: Дата окончания пересчета в формате YYYY-MM-DD (по умолчанию: сегодня)
        """
        try:
            # Получаем параметры
            if hasattr(request, 'data') and request.data:
                start_date_str = request.data.get('start_date')
                end_date_str = request.data.get('end_date')
            else:
                start_date_str = request.query_params.get('start_date')
                end_date_str = request.query_params.get('end_date')
            
            # Парсим даты
            start_date = None
            end_date = None
            
            if start_date_str:
                try:
                    if ' ' in start_date_str or 'T' in start_date_str:
                        start_date = datetime.strptime(start_date_str.replace('T', ' '), "%Y-%m-%d %H:%M:%S")
                    else:
                        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                except ValueError as e:
                    return JsonResponse({
                        "status": "error",
                        "message": f"Неверный формат start_date: {e}"
                    }, status=400)
            
            if end_date_str:
                try:
                    if ' ' in end_date_str or 'T' in end_date_str:
                        end_date = datetime.strptime(end_date_str.replace('T', ' '), "%Y-%m-%d %H:%M:%S")
                    else:
                        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                except ValueError as e:
                    return JsonResponse({
                        "status": "error",
                        "message": f"Неверный формат end_date: {e}"
                    }, status=400)
            
            # Если даты не указаны, используем значения по умолчанию: с 1 декабря по сегодня
            if not start_date:
                today = timezone.now().date()
                start_date = datetime.combine(datetime(today.year, 12, 1).date(), datetime.min.time())
                start_date = timezone.make_aware(start_date)
                logger.info(f"Using default start_date: {start_date}")
            
            if not end_date:
                end_date = timezone.now()
                logger.info(f"Using default end_date: {end_date}")
            
            # Делаем даты aware, если нужно
            if start_date and timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date)
            if end_date and timezone.is_naive(end_date):
                end_date = timezone.make_aware(end_date)
            
            # Запускаем пересчет в отдельном потоке, чтобы не блокировать ответ
            import threading
            
            def run_full_recalculate():
                try:
                    logger.info(f"Начинаем полный пересчет с {start_date} по {end_date}")
                    
                    # Шаг 1: Пересчитываем EntryExit из CameraEvent
                    logger.info("Шаг 1: Пересчет EntryExit из CameraEvent...")
                    result_entries = recalculate_entries_exits(start_date=start_date, end_date=end_date)
                    logger.info(f"EntryExit пересчет завершен: создано={result_entries.get('created', 0)}, обновлено={result_entries.get('updated', 0)}")
                    
                    # Шаг 2: Пересчитываем статистику посещаемости
                    logger.info("Шаг 2: Пересчет статистики посещаемости...")
                    import sys
                    from pathlib import Path
                    
                    project_root = Path(__file__).resolve().parent.parent
                    if str(project_root) not in sys.path:
                        sys.path.insert(0, str(project_root))
                    
                    from recalculate_attendance_stats import recalculate_attendance_stats
                    
                    # Используем только дату (без времени) для пересчета статистики
                    start_date_only = start_date.date() if hasattr(start_date, 'date') else start_date
                    recalculate_attendance_stats(start_date=start_date_only)
                    
                    logger.info("Полный пересчет завершен успешно!")
                except Exception as e:
                    logger.error(f"Ошибка при полном пересчете: {e}", exc_info=True)
            
            thread = threading.Thread(target=run_full_recalculate, daemon=True)
            thread.start()
            
            return JsonResponse({
                "status": "success",
                "message": "Полный пересчет запущен в фоновом режиме",
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
                "note": "Пересчет включает: 1) EntryExit из CameraEvent, 2) Статистику посещаемости"
            })
        except Exception as e:
            logger.error(f"Ошибка при запуске полного пересчета: {e}", exc_info=True)
            return JsonResponse({
                "status": "error",
                "message": str(e)
            }, status=500)
class AttendanceStatsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet для экспорта статистики посещаемости по подразделениям.
    """
    queryset = Employee.objects.none()  # Не используется для стандартных операций
    permission_classes = [AllowAny]
    
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Экспорт статистики посещаемости по подразделениям в Excel.
        Фильтрует сотрудников по department_id (может быть несколько) и экспортирует их данные.
        
        Параметры:
        - department_id - ID подразделения (можно указать несколько раз)
        - start_date - начальная дата (формат: YYYY-MM-DD)
        - end_date - конечная дата (формат: YYYY-MM-DD)
        """
        # Получаем параметры
        department_ids = request.query_params.getlist("department_id")
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")
        
        # Валидация
        if not department_ids:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ошибка"
            ws.append(["Не указаны ID подразделений. Используйте параметр department_id."])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="error.xlsx"'
            return response
        
        return self._export_excel_by_department_ids(department_ids, start_date_str, end_date_str)
    
    def _export_excel_by_department_ids(self, department_ids, start_date_str, end_date_str):
        """
        Экспорт данных по ID подразделений с использованием SQL запросов.
        """
        from .sql_reports import generate_comprehensive_attendance_report_sql
        from .utils import get_excluded_hikvision_ids
        from django.db.models import Q
        
        # Получаем исключаемые ID
        excluded_hikvision_ids = get_excluded_hikvision_ids()
        
        # Функция для получения всех дочерних подразделений
        def get_all_children(dept_obj):
            children = [dept_obj.id]
            for child in dept_obj.children.all():
                children.extend(get_all_children(child))
            return children
        
        # Собираем все ID подразделений (включая дочерние)
        all_department_ids = []
        for dept_id in department_ids:
            try:
                dept_id_int = int(dept_id)
                dept = Department.objects.filter(id=dept_id_int).first()
                if dept:
                    all_department_ids.extend(get_all_children(dept))
            except (ValueError, TypeError):
                continue
        
        # Убираем дубликаты
        all_department_ids = list(set(all_department_ids))
        
        if not all_department_ids:
            # Если не найдено подразделений, возвращаем пустой Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Нет данных"
            ws.append(["Не найдено подразделений по указанным ID"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="no_data.xlsx"'
            return response
        
        # Ищем сотрудников по подразделениям
        employees_query = Employee.objects.exclude(
            hikvision_id__in=excluded_hikvision_ids
        ).filter(
            hikvision_id__isnull=False,
            department_id__in=all_department_ids
        )
        
        employees_to_export = list(
            employees_query.select_related('department').prefetch_related('work_schedules').distinct()
        )
        
        if not employees_to_export:
            # Если не найдено сотрудников, возвращаем пустой Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Нет данных"
            ws.append(["Не найдено сотрудников в указанных подразделениях"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="no_data.xlsx"'
            return response
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        # Удаляем дефолтный лист
        if wb.worksheets:
            wb.remove(wb.worksheets[0])
        
        # Для каждого сотрудника создаем отдельный лист
        for employee in employees_to_export:
            emp_hikvision_id = employee.hikvision_id
            
            # Получаем данные для этого сотрудника
            results, start_date_obj, end_date_obj = generate_comprehensive_attendance_report_sql(
                hikvision_id=emp_hikvision_id,
                start_date=start_date_str,
                end_date=end_date_str,
                device_name=None,
                excluded_hikvision_ids=excluded_hikvision_ids
            )
            
            # Создаем лист для сотрудника
            # Ограничиваем длину имени листа (Excel ограничение - 31 символ)
            sheet_name = (employee.name or f"ID_{emp_hikvision_id}")[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Используем метод из EntryExitViewSet для заполнения листа
            # Создаем временный экземпляр для вызова метода
            entry_exit_viewset = EntryExitViewSet()
            entry_exit_viewset._fill_employee_sheet(ws, employee, results, start_date_obj, end_date_obj)
        
        # Если нет ни одного листа, создаем пустой
        if len(wb.worksheets) == 0:
            ws = wb.create_sheet(title="Нет данных")
            ws.append(["Не найдено данных"])
        
        # Сохраняем файл
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Генерируем имя файла
        if len(employees_to_export) == 1:
            emp = employees_to_export[0]
            emp_name = re.sub(r'[<>:"/\\|?*]', '_', emp.name or f"ID_{emp.hikvision_id}")
            emp_name = emp_name.replace(' ', '_')
            emp_name = re.sub(r'_+', '_', emp_name)
            if start_date_str and end_date_str:
                start_date_obj = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date_obj = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                start_date_str_formatted = start_date_obj.strftime('%d-%m-%Y')
                end_date_str_formatted = end_date_obj.strftime('%d-%m-%Y')
                filename = f"{emp_name}_с_{start_date_str_formatted}_по_{end_date_str_formatted}.xlsx"
            else:
                filename = f"{emp_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"отчет_по_подразделениям_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = FileResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        from urllib.parse import quote
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(filename)}'
        return response


# DepartmentViewSet вынесен в viewsets/department.py
# Импортируется выше