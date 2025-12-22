"""
Утилита для импорта и экспорта сотрудников из/в Excel файлы.
"""
import os
import django
from openpyxl import load_workbook, Workbook
from django.db import transaction

# Настройка Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'hikvision_project.settings')
django.setup()

from .models import Employee, Department, WorkSchedule, CameraEvent


def clean_id(id_str):
    """Удаляет ведущие нули из ID."""
    if not id_str:
        return None
    s = str(id_str).strip()
    if s.replace('0', '') == '':
        return "0"
    return s.lstrip('0') or "0"


def parse_time_string(time_str):
    """
    Парсит строку времени в формате 'HH:MM' или 'HH:MM-HH:MM'.
    Возвращает tuple (start_time, end_time) или None.
    """
    if not time_str:
        return None, None
    
    time_str = str(time_str).strip()
    
    # Если формат 'HH:MM-HH:MM'
    if '-' in time_str and ':' in time_str:
        parts = time_str.split('-')
        if len(parts) == 2:
            try:
                from datetime import time as dt_time
                start_parts = parts[0].strip().split(':')
                end_parts = parts[1].strip().split(':')
                start_time = dt_time(int(start_parts[0]), int(start_parts[1]))
                end_time = dt_time(int(end_parts[0]), int(end_parts[1]))
                return start_time, end_time
            except (ValueError, IndexError):
                pass
    
    return None, None


def get_or_create_department(dept_name):
    """
    Получает или создает подразделение по имени.
    Поддерживает иерархию через ' > '.
    """
    if not dept_name:
        return None
    
    dept_name = str(dept_name).strip()
    if not dept_name:
        return None
    
    # Проверяем, есть ли иерархия
    if ' > ' in dept_name:
        parts = dept_name.split(' > ')
        parent = None
        for part in parts[:-1]:
            part = part.strip()
            if part:
                parent, _ = Department.objects.get_or_create(
                    name=part,
                    parent=parent,
                    defaults={'name': part, 'parent': parent}
                )
        
        dept_name = parts[-1].strip()
        if dept_name:
            department, _ = Department.objects.get_or_create(
                name=dept_name,
                parent=parent,
                defaults={'name': dept_name, 'parent': parent}
            )
            return department
    else:
        # Обычное подразделение без иерархии
        department, _ = Department.objects.get_or_create(
            name=dept_name,
            defaults={'name': dept_name}
        )
        return department
    
    return None


def parse_schedule_type(schedule_type_str):
    """
    Парсит тип графика из строки.
    Возвращает один из: 'regular', 'floating', 'round_the_clock'
    """
    if not schedule_type_str:
        return 'regular'
    
    schedule_type_str = str(schedule_type_str).strip().lower()
    
    if 'плавающий' in schedule_type_str or 'floating' in schedule_type_str:
        return 'floating'
    elif 'круглосуточн' in schedule_type_str or 'round' in schedule_type_str or '24' in schedule_type_str:
        return 'round_the_clock'
    else:
        return 'regular'


def import_employees_from_excel(file_path, update_existing=True):
    """
    Импортирует сотрудников из Excel файла.
    
    Параметры:
        file_path (str): Путь к Excel файлу
        update_existing (bool): Если True, обновляет существующих сотрудников. 
                                Если False, пропускает существующих.
    
    Возвращает:
        dict: Словарь с результатами импорта:
            - 'success': количество успешно импортированных
            - 'updated': количество обновленных
            - 'created': количество созданных
            - 'errors': список ошибок
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    results = {
        'success': 0,
        'updated': 0,
        'created': 0,
        'errors': []
    }
    
    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
        ws = wb.active
        
        # Пропускаем заголовок (первую строку)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        with transaction.atomic():
            for row_num, row in enumerate(rows, start=2):
                try:
                    # Проверяем, есть ли данные в строке
                    if not row or not any(row):
                        continue
                    
                    # Извлекаем данные
                    employee_id_raw = row[0] if len(row) > 0 else None
                    name = row[1] if len(row) > 1 else None
                    department_name = row[2] if len(row) > 2 else None
                    position = row[3] if len(row) > 3 else None
                    schedule_type_str = row[4] if len(row) > 4 else None
                    schedule_str = row[5] if len(row) > 5 else None
                    allowed_late_minutes = row[6] if len(row) > 6 else 0
                    allowed_early_leave_minutes = row[7] if len(row) > 7 else 0
                    
                    # Валидация обязательных полей
                    if not employee_id_raw:
                        results['errors'].append(
                            f"Строка {row_num}: Отсутствует Employee ID"
                        )
                        continue
                    
                    if not name:
                        results['errors'].append(
                            f"Строка {row_num}: Отсутствует имя сотрудника"
                        )
                        continue
                    
                    # Очищаем и нормализуем ID
                    employee_id = clean_id(employee_id_raw)
                    
                    # Получаем или создаем подразделение
                    department = None
                    if department_name:
                        department = get_or_create_department(department_name)
                    
                    # Обрабатываем имя
                    name = str(name).strip().replace('\n', ' ').replace('\r', ' ')
                    import re
                    name = re.sub(r'\s+', ' ', name)
                    
                    # Получаем или создаем сотрудника
                    employee, created = Employee.objects.get_or_create(
                        hikvision_id=employee_id,
                        defaults={
                            'name': name,
                            'department': department,
                            'position': str(position).strip() if position else None,
                        }
                    )
                    
                    # Обновляем существующего сотрудника, если нужно
                    if not created and update_existing:
                        employee.name = name
                        employee.department = department
                        if position:
                            employee.position = str(position).strip()
                        employee.save()
                        results['updated'] += 1
                    elif created:
                        results['created'] += 1
                    
                    # Обрабатываем график работы
                    # Если есть данные о графике в Excel, заменяем все существующие графики новым
                    if schedule_type_str or schedule_str:
                        schedule_type = parse_schedule_type(schedule_type_str)
                        
                        # Преобразуем допустимые минуты в числа
                        try:
                            allowed_late = int(allowed_late_minutes) if allowed_late_minutes else 0
                        except (ValueError, TypeError):
                            allowed_late = 0
                        
                        try:
                            allowed_early = int(allowed_early_leave_minutes) if allowed_early_leave_minutes else 0
                        except (ValueError, TypeError):
                            allowed_early = 0
                        
                        # Парсим график
                        start_time = None
                        end_time = None
                        description = None
                        
                        if schedule_str:
                            schedule_str = str(schedule_str).strip()
                            if schedule_type == 'regular' and ('-' in schedule_str or ':' in schedule_str):
                                start_time, end_time = parse_time_string(schedule_str)
                                if start_time and end_time:
                                    description = f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"
                            else:
                                description = schedule_str
                        
                        # Удаляем все существующие графики сотрудника и создаем новый
                        # Это гарантирует замену графика новыми данными при несовпадении
                        WorkSchedule.objects.filter(employee=employee).delete()
                        
                        # Создаем новый график с данными из Excel
                        WorkSchedule.objects.create(
                            employee=employee,
                            schedule_type=schedule_type,
                            start_time=start_time,
                            end_time=end_time,
                            description=description,
                            allowed_late_minutes=allowed_late,
                            allowed_early_leave_minutes=allowed_early,
                        )
                    
                    results['success'] += 1
                    
                except Exception as e:
                    error_msg = f"Строка {row_num}: {str(e)}"
                    results['errors'].append(error_msg)
                    continue
        
        wb.close()
        
    except Exception as e:
        results['errors'].append(f"Критическая ошибка при чтении файла: {str(e)}")
    
    return results


def extract_employee_name_from_event(camera_event):
    """
    Извлекает имя сотрудника из raw_data события камеры.
    """
    if not camera_event or not camera_event.raw_data or not isinstance(camera_event.raw_data, dict):
        return None
    
    outer_event = camera_event.raw_data.get("AccessControllerEvent", {})
    # Проверяем вложенную структуру
    if isinstance(outer_event, dict) and "AccessControllerEvent" in outer_event:
        access_event = outer_event["AccessControllerEvent"]
    else:
        access_event = outer_event if isinstance(outer_event, dict) else {}
    
    return (
        access_event.get("employeeName") or
        access_event.get("name") or
        access_event.get("employeeNameString") or
        None
    )


def export_employees_to_excel(file_path, department_filter=None):
    """
    Экспортирует сотрудников из базы данных в Excel файл.
    Включает также сотрудников из CameraEvent, которых еще нет в Employee.
    
    Параметры:
        file_path (str): Путь к Excel файлу для сохранения
        department_filter (Department, optional): Если указан, экспортирует только сотрудников этого подразделения
    
    Возвращает:
        dict: Словарь с результатами экспорта:
            - 'success': количество экспортированных сотрудников
            - 'from_employee_table': количество из таблицы Employee
            - 'from_camera_events': количество из CameraEvent
            - 'errors': список ошибок
    """
    results = {
        'success': 0,
        'from_employee_table': 0,
        'from_camera_events': 0,
        'errors': []
    }
    
    try:
        # Создаем новый Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"
        
        # Добавляем заголовки
        headers = [
            'Employee ID',
            'Имя',
            'Подразделение',
            'Должность',
            'Тип графика',
            'График',
            'Допустимое опоздание',
            'Допустимый ранний уход'
        ]
        ws.append(headers)
        
        # Форматируем заголовки (жирный шрифт)
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Получаем существующих сотрудников
        if department_filter:
            employees = Employee.objects.filter(department=department_filter).select_related('department').prefetch_related('work_schedules')
        else:
            employees = Employee.objects.select_related('department').prefetch_related('work_schedules')
        
        # Функция для числовой сортировки ID
        def get_id_sort_key(hikvision_id):
            """Возвращает ключ для числовой сортировки ID."""
            normalized_id = clean_id(hikvision_id)
            try:
                # Пытаемся преобразовать в число
                return int(normalized_id)
            except (ValueError, TypeError):
                # Если не число, возвращаем большую константу, чтобы такие ID шли в конце
                return float('inf')
        
        # Сортируем сотрудников по числовому значению ID
        employees_list = list(employees)
        employees_list.sort(key=lambda emp: (get_id_sort_key(emp.hikvision_id), emp.hikvision_id))
        
        # Множество ID существующих сотрудников для быстрой проверки
        existing_employee_ids = set()
        
        # Экспортируем данные из таблицы Employee
        for employee in employees_list:
            try:
                existing_employee_ids.add(clean_id(employee.hikvision_id))
                
                # Получаем подразделение (с полным путем, если есть иерархия)
                department_name = ""
                if employee.department:
                    department_name = employee.department.get_full_path()
                elif employee.department_old:
                    department_name = employee.department_old
                
                # Получаем график работы (берем первый, если их несколько)
                work_schedule = employee.work_schedules.first()
                
                schedule_type_display = ""
                schedule_display = ""
                allowed_late = 0
                allowed_early = 0
                
                if work_schedule:
                    # Тип графика (читаемое название)
                    schedule_type_display = work_schedule.get_schedule_type_display()
                    
                    # Описание графика
                    schedule_display = work_schedule.get_schedule_display()
                    
                    # Допустимые отклонения
                    allowed_late = work_schedule.allowed_late_minutes
                    allowed_early = work_schedule.allowed_early_leave_minutes
                
                # Добавляем строку данных
                row_data = [
                    employee.hikvision_id,
                    employee.name,
                    department_name,
                    employee.position or "",
                    schedule_type_display,
                    schedule_display,
                    allowed_late,
                    allowed_early,
                ]
                ws.append(row_data)
                
                results['success'] += 1
                results['from_employee_table'] += 1
                
            except Exception as e:
                error_msg = f"Ошибка при экспорте сотрудника {employee.hikvision_id}: {str(e)}"
                results['errors'].append(error_msg)
                continue
        
        # Теперь добавляем сотрудников из CameraEvent, которых нет в Employee
        # Получаем все события, отсортированные по времени (от новых к старым)
        # Это позволит получить последнее событие для каждого ID в одном запросе
        all_events = CameraEvent.objects.exclude(
            hikvision_id__isnull=True
        ).exclude(
            hikvision_id=''
        ).order_by('hikvision_id', '-event_time', '-created_at')
        
        # Словарь для хранения нормализованных ID и их последних событий
        # Ключ - нормализованный ID, значение - последнее событие
        normalized_events = {}
        
        # Обрабатываем события - берем только первое (последнее по времени) для каждого оригинального ID
        processed_original_ids = set()
        
        for event in all_events:
            try:
                original_id = event.hikvision_id
                
                # Пропускаем, если уже обработали этот оригинальный ID
                if original_id in processed_original_ids:
                    continue
                
                # Нормализуем ID
                hikvision_id = clean_id(original_id)
                
                # Пропускаем, если сотрудник уже есть в Employee
                if hikvision_id in existing_employee_ids:
                    processed_original_ids.add(original_id)
                    continue
                
                # Если этот нормализованный ID уже обработан, пропускаем
                # (это предотвратит дубликаты для случаев типа "001" и "1")
                if hikvision_id in normalized_events:
                    processed_original_ids.add(original_id)
                    continue
                
                # Сохраняем последнее событие для этого нормализованного ID
                normalized_events[hikvision_id] = event
                processed_original_ids.add(original_id)
                
            except Exception as e:
                error_msg = f"Ошибка при обработке события {event.id}: {str(e)}"
                results['errors'].append(error_msg)
                continue
        
        # Сортируем сотрудников из CameraEvent по числовому значению ID
        def get_event_id_sort_key(item):
            """Возвращает ключ для числовой сортировки ID из событий."""
            hikvision_id, _ = item
            return get_id_sort_key(hikvision_id)
        
        sorted_events = sorted(normalized_events.items(), key=get_event_id_sort_key)
        
        # Теперь добавляем сотрудников из CameraEvent в Excel
        for hikvision_id, event in sorted_events:
            try:
                # Извлекаем имя из raw_data
                employee_name = extract_employee_name_from_event(event)
                
                # Если имени нет, используем ID как имя
                if not employee_name:
                    employee_name = f"Сотрудник {hikvision_id}"
                
                # Добавляем строку данных (без подразделения, должности и графика)
                row_data = [
                    hikvision_id,
                    employee_name,
                    "",  # Подразделение
                    "",  # Должность
                    "",  # Тип графика
                    "",  # График
                    0,   # Допустимое опоздание
                    0,   # Допустимый ранний уход
                ]
                ws.append(row_data)
                
                results['success'] += 1
                results['from_camera_events'] += 1
                
            except Exception as e:
                error_msg = f"Ошибка при экспорте сотрудника из CameraEvent (ID: {hikvision_id}): {str(e)}"
                results['errors'].append(error_msg)
                continue
        
        # Настраиваем ширину колонок
        column_widths = [15, 30, 30, 25, 20, 25, 20, 25]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Сохраняем файл
        wb.save(file_path)
        
    except Exception as e:
        results['errors'].append(f"Критическая ошибка при экспорте: {str(e)}")
    
    return results


if __name__ == '__main__':
    """
    Пример использования:
    python -m camera_events.import_employees
    """
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python -m camera_events.import_employees <путь_к_excel_файлу>")
        print("Пример: python -m camera_events.import_employees employee_import_template.xlsx")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    try:
        results = import_employees_from_excel(file_path, update_existing=True)
        
        print(f"\n=== Результаты импорта ===")
        print(f"Успешно обработано: {results['success']}")
        print(f"Создано новых: {results['created']}")
        print(f"Обновлено: {results['updated']}")
        
        if results['errors']:
            print(f"\nОшибки ({len(results['errors'])}):")
            for error in results['errors']:
                print(f"  - {error}")
        else:
            print("\nОшибок не обнаружено!")
            
    except Exception as e:
        print(f"Ошибка при импорте: {str(e)}")
        sys.exit(1)

